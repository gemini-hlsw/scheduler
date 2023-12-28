# Copyright (c) 2016-2022 Association of Universities for Research in Astronomy, Inc. (AURA)
# For license information see LICENSE or https://opensource.org/licenses/BSD-3-Clause

import os
import csv
from copy import copy
from datetime import date, datetime, timedelta
from typing import Dict, List, Set, Tuple, Union, Final
from io import BytesIO, StringIO

from lucupy.minimodel import Site, ALL_SITES, Resource
from lucupy.helpers import str_to_bool
import gelidum
import requests
from openpyxl import load_workbook

from definitions import ROOT_DIR
from .filters import *
from .night_resource_configuration import NightConfiguration
from .google_drive_downloader import GoogleDriveDownloader
from scheduler.services.abstract import ExternalService
from scheduler.services import logger_factory

logger = logger_factory.create_logger(__name__)


class ResourceService(ExternalService):
    """
    This is to avoid recreating repetitive resources.
    When we first get a resource ID string, create a Resource for it and store it here.
    Then fetch the Resources from here if they exist, and if they do not, then create a new one as per
    the lookup_resource method.
    """
    # These are the converters from the OCS FPU names to the ITCD FPU representations.
    # For example, the ODB query extractor would return:
    #    * 'IFU 2 Slits'
    # which we would want to convert to:
    #    * 'IFU-2'
    # since these are the FPU names used in the GMOS[NS]-FPUr######.txt files.
    _gmosn_ifu_dict: Dict[str, str] = gelidum.freeze({
        'IFU 2 Slits': 'IFU-2',
        'IFU Left Slit (blue)': 'IFU-B',
        'IFU Right Slit (red)': 'IFU-R',
        'Longslit 0.25 arcsec': '0.25arcsec',
        'Longslit 0.50 arcsec': '0.5arcsec',
        'Longslit 0.75 arcsec': '0.75arcsec',
        'Longslit 1.00 arcsec': '1.0arcsec',
        'Longslit 1.5 arcsec': '1.5arcsec',
        'Longslit 2.00 arcsec': '2.0arcsec',
        'Longslit 5.00 arcsec': '5.0arcsec',
        'N and S 0.50 arcsec': 'NS0.5arcsec',
        'N and S 0.75 arcsec': 'NS0.75arcsec',
        'N and S 1.00 arcsec': 'NS1.0arcsec',
        'N and S 1.50 arcsec': 'NS1.5arcsec',
        'N and S 2.00 arcsec': 'NS2.0arcsec',
        'focus_array_new': 'focus_array_new'
    })

    _gmoss_ifu_dict: Dict[str, str] = gelidum.freeze({**_gmosn_ifu_dict, **{
        'IFU N and S 2 Slits': 'IFU-NS-2',
        'IFU N and S Left Slit (blue)': 'IFU-NS-B',
        'IFU N and S Right Slit (red)': 'IFU-NS-R',
        # TODO: Not in OCS?
        'PinholeC': 'PinholeC'
    }})

    # Constants for converting MDF to barcodes.
    _instd = {'GMOS': '1', 'GMOS-N': '1', 'GMOS-S': '1', 'Flamingos2': '3'}
    _semd = {'A': '0', 'B': '1'}
    _progd = {'Q': '0', 'C': '1', 'L': '2', 'F': '3', 'S': '8', 'D': '9'}

    def __init__(self, sites: FrozenSet[Site] = ALL_SITES):
        self._all_resources: Dict[str, Resource] = {}
        self._sites = sites
        self._path = os.path.join(ROOT_DIR, 'scheduler', 'services', 'resource', 'data')

        # The map from site and date to the set of resources.
        self._resources: Dict[Site, Dict[date, Set[Resource]]] = {site: {} for site in self._sites}

        # Mapping from ITCD FPUs to barcodes. The mapping is site-dependent.
        # The ODB program extractor produces long versions of these names that must be run through the
        # OcsFpuConverter to get the ITCD FPU names.
        self._itcd_fpu_to_barcode: Dict[Site, Dict[str, Resource]] = {site: {} for site in self._sites}

        # Earliest and latest dates per site.
        self._earliest_date_per_site: Dict[Site, date] = {site: date.max for site in self._sites}
        self._latest_date_per_site: Dict[Site, date] = {site: date.min for site in self._sites}

        # Determines which nights are blocked.
        self._blocked: Dict[Site, Set[date]] = {site: set() for site in self._sites}

        # Determines whether a night is a part of a laser run.
        self._lgs: Dict[Site, Dict[date, bool]] = {site: {} for site in self._sites}

        # Determines whether ToOs are accepted on a given night.
        self._too: Dict[Site, Dict[date, bool]] = {site: {} for site in self._sites}

        # Filters to apply to a night. We add the ResourceFilters at the end after all the resources
        # have been processed.
        self._positive_filters: Dict[Site, Dict[date, Set[AbstractFilter]]] = {site: {} for site in self._sites}
        self._negative_filters: Dict[Site, Dict[date, Set[AbstractFilter]]] = {site: {} for site in self._sites}

        # The final combined filters for each night.
        self._filters: Dict[Site, Dict[date, AbstractFilter]] = {site: {} for site in self._sites}

        # The final output from this class: the configuration per night.
        self._night_configurations: Dict[Site, Dict[date, NightConfiguration]] = {site: {} for site in self._sites}

    def _mdf_to_barcode(self, mdfname: str, inst: str) -> Optional[Resource]:
        """Legacy MOS mask barcode convention"""
        barcode = None
        if inst in ResourceService._instd.keys():
            # Collect the components of the string from the MDF name.
            inst_id = ResourceService._instd[inst]
            sem_id = ResourceService._semd[mdfname[6]]
            progtype_id = ResourceService._progd[mdfname[7]]
            barcode = f'{inst_id}{sem_id}{progtype_id}{mdfname[-6:-3]}{mdfname[-2:]}'
        return self.lookup_resource(barcode)

    def _itcd_fpu_to_barcode_parser(self, r: List[str], site: Site) -> Set[str]:
        return {self._itcd_fpu_to_barcode[site][r[0].strip()].id} | {i.strip() for i in r[1:]}

    def _convert_fpu_to_itcd_name(self, site: Site, fpu_name: str) -> Optional[str]:
        """
        Convert a long FPU name into its ITCD name, if it exists.
        """
        if Site.GN in self._sites and site == Site.GN:
            return self._gmosn_ifu_dict.get(fpu_name)
        if Site.GS in self._sites and site == Site.GS:
            return self._gmoss_ifu_dict.get(fpu_name)
        return None

    def lookup_resource(self, resource_id: str) -> Optional[Resource]:
        """
        Function to perform Resource caching and minimize the number of Resource objects by attempting to reuse
        Resource objects with the same ID.

        If resource_id evaluates to False, return None.
        Otherwise, check if a Resource with id already exists.
        If it does, return it.
        If not, create it, add it to the map of all Resources, and then return it.

        Note that even if multiple objects do exist with the same ID, they will be considered equal by the
        Resource equality comparator.
        """
        # The Resource constructor raises an exception for id None or containing any capitalization of "none".
        if not resource_id:
            return None
        if resource_id not in self._all_resources:
            self._all_resources[resource_id] = Resource(id=resource_id)
        return self._all_resources[resource_id]

    def date_range_for_site(self, site: Site) -> Tuple[date, date]:
        """
        Return the date range (inclusive) for which we have resource data for a site.
        """
        if site not in self._sites:
            raise ValueError(f'Request for resource dates for illegal site: {site.name}')
        return self._earliest_date_per_site[site], self._latest_date_per_site[site]

    def get_night_configuration(self, site: Site, local_date: date) -> NightConfiguration:
        """
        Returns the NightConfiguration object for the site for the given local date,
        which contains the filters and resources for the night.
        """
        if site not in self._sites:
            raise ValueError(f'Request for night configuration for illegal site: {site.name}')
        if local_date < self._earliest_date_per_site[site] or local_date > self._latest_date_per_site[site]:
            raise ValueError(f'Request for night configuration for site {site.name} for illegal date: {local_date}')
        return self._night_configurations[site][local_date]

    def get_resources(self, site: Site, night_date: date) -> FrozenSet[Resource]:
        """
        For a site and a local date, return the set of available resources.
        The date is currently the truncation to day of the astropy Time objects in the time_grid, which are in UTC
           and have times of 8:00 am, so to get local dates, in the Collector we subtract one day.
        If the date falls before any resource data for the site, return the empty set.
        If the date falls after any resource data for the site, return the last resource set.
        """
        if site not in self._sites:
            raise ValueError(f'Request for resources for illegal site: {site.name}')

        # If the date is before the first date or after the last date, return the empty set.
        if night_date < self._earliest_date_per_site[site] or night_date > self._latest_date_per_site[site]:
            return frozenset()

        return frozenset(self._resources[site][night_date])

    def fpu_to_barcode(self, site: Site, fpu_name: str, instrument: str) -> Optional[Resource]:
        """
        Convert a long FPU name into the barcode, if it exists.
        """
        barcode = None
        itcd_fpu_name = self._convert_fpu_to_itcd_name(site, fpu_name)
        # print(f'fpu_to_barcode {fpu_name} {instrument} {itcd_fpu_name}')
        if itcd_fpu_name:
            barcode = self._itcd_fpu_to_barcode[site].get(itcd_fpu_name)
        elif fpu_name.startswith('G'):
            barcode = self._mdf_to_barcode(fpu_name, inst=instrument)
        # print(f'\t barcode {barcode}')
        return barcode


class FileBasedResourceService(ResourceService):
    """
    This is the base for both OCS and File-uploaded services
    """
    # Definition of a day to not have to redeclare constantly.
    _day: Final[timedelta] = timedelta(days=1)
    # Statuses of instruments and WFS.
    _SCIENCE: Final[str] = 'SCIENCE'
    _ENGINEERING: Final[str] = 'ENGINEERING'
    _NOT_AVAILABLE: Final[str] = 'NOT AVAILABLE'
    _CALIBRATION: Final[str] = 'CALIBRATION'

    # The Google ID of the telescope configuration file.
    _SITE_CONFIG_GOOGLE_ID: Final[str] = '1QRalQNEaX-bcyrPG6mfKnv01JVMaGHwy'

    # Name of the spreadsheet file containing telescope configurations.
    _SITE_CONFIG_FILE: Final[str] = 'telescope_schedules.xlsx'

    def _load_fpu_to_barcodes(self, site: Site, name: str) -> None:
        """
        FPUs at each site map to a unique barcode as defined in the files:
            * gmos[ns]_fpu_barcode.txt
        These are site-dependent values.
        """
        with open(os.path.join(self._path, name)) as f:
            for row in f:
                fpu, barcode = row.split()

                # Only map if the FPU is a resource.
                if fpu is not None:
                    self._itcd_fpu_to_barcode[site][fpu] = self.lookup_resource(barcode)

    def _load_csv(self, site: Site,
                  c: Callable[[List[str], Site], Set[str]],
                  data_source: Union[str, BytesIO]) -> None:
        """
        Process a CSV file as a table, where:

        1. The first entry is a date in YYYY-mm-dd format
        2. The remaining entries are resources available on that date to the following date in the file.

        If a date is missing from the CSV file, copy the data from the previously defined date through to just before
        the new date.
        """

        def _process_file(f) -> None:
            reader = csv.reader(f, delimiter=',')
            prev_row_date: Optional[date] = None

            for row in reader:
                # Get rid of the byte-order marker, which causes datetime.strptime to gail.
                row = [col.replace('\ufeff', '') for col in row]
                row_date = datetime.strptime(row[0].strip(), '%Y-%m-%d').date()

                # Fill in any gaps by copying prev_row_date until we reach one less than row_date.
                if prev_row_date is not None:
                    missing_row_date = prev_row_date + FileBasedResourceService._day
                    while missing_row_date < row_date:
                        # Make sure there is an entry and append to it to avoid overwriting anything already present.
                        date_set = self._resources[site].setdefault(missing_row_date, set())
                        self._resources[site][missing_row_date] = date_set | copy(self._resources[site][prev_row_date])
                        missing_row_date += FileBasedResourceService._day

                # Get or create date_set for the date, and append new resources from table, ignoring blank entries.
                date_set = self._resources[site].setdefault(row_date, set())
                new_entries = {self.lookup_resource(r) for r in c(row[1:], site) if r}
                self._resources[site][row_date] = date_set | new_entries

                # Advance the previous row date where data was defined.
                prev_row_date = row_date

        if isinstance(data_source, str):
            with open(os.path.join(self._path, data_source)) as f:
                _process_file(f)
        else:
            data_str = data_source.getvalue().decode()
            f = StringIO(data_str)
            _process_file(f)

    @staticmethod
    def _split_program_ids(text: str) -> Set[ProgramID]:
        """
        Given a text list that should contain comma-separated program IDs, separate them
        and return the set of them.
        """
        return set(ProgramID(prog_id.strip()) for prog_id in text.strip().split(','))

    @staticmethod
    def _add_dates_to_dict(prog_dict: Dict[ProgramID, Set[date]], text: str, add_date: date) -> None:
        """
        Given a program dictionary, parse the text following it into program IDs and add the
        specified date to each set associated with each program ID.
        """
        for prog_id in FileBasedResourceService._split_program_ids(text):
            date_set = prog_dict.setdefault(prog_id, set())
            date_set.add(add_date)

    def _load_spreadsheet(self, file_source: Union[str, BytesIO],
                          from_gdrive: bool = False) -> None:
        """
        Process an Excel spreadsheet containing instrument, mode, and LGS information.

        The Excel spreadsheets have information available for every date, so we do not have to concern ourselves
        as in the _load_csv file above.
        """

        def none_to_str(value) -> str:
            return '' if value is None else value

        if not file_source:
            raise ValueError('file_source cannot be empty')

        # filename = os.path.join(self._path, OcsResourceService._SITE_CONFIG_FILE)
        if from_gdrive:
            logger.info('Retrieving site configuration file from Google Drive...')
            try:
                GoogleDriveDownloader.download_file_from_google_drive(file_id=FileBasedResourceService._SITE_CONFIG_GOOGLE_ID,
                                                                      overwrite=True,
                                                                      dest_path=file_source)
            except requests.RequestException:
                logger.warning('Could not retrieve site configuration file from Google Drive.')

            if not os.path.exists(file_source):
                raise FileNotFoundError(
                    f'No site configuration data available for {__class__.__name__} at: {file_source}')

        workbook = load_workbook(filename=file_source,
                                 read_only=True,
                                 data_only=True)
        for site in self._sites:
            try:
                sheet = workbook[site.name]
            except KeyError:
                # Make the KeyError more clear.
                raise KeyError(f'{__class__.__name__}: No tab for {site.name} in the '
                               'telescope schedule configuration spreadsheet.')

            # A set consisting of the dates that are not blocked.
            dates: Set[date] = set()

            # We need to determine what programs are prohibited from what nights.
            # PV programs can only be done on PV nights or later, so we store the night they first appear.
            # They also receive boosted pri
            pv_programs: Dict[ProgramID, Set[date]] = {}

            # Classical programs can only be done on specifically designated nights.
            classical_programs: Dict[ProgramID, Set[date]] = {}

            # Programs that get a boost on dates.
            score_boost: Dict[ProgramID, Set[date]] = {}

            # Partner blocks are when the mode indicates that the night is restricted to a partner.
            partner_blocks: Dict[date, TimeAccountingCode] = {}

            # Instrument runs, where observations associated with certain instruments get a boost.
            instrument_run: Dict[Resource, Set[date]] = {}

            # Process the header.
            # We need to make mappings from all the instrument names to their column number.
            # We also need the names of the WFS columns. These are reversed because we want to go from column idx to
            #   resource.
            instrument_column_mapping: Dict[str, int] = {}
            wfs_columns: Dict[int, str] = {}

            column_idx = 11

            # sheet.max_column is returning None, probably because of missing data in PWFS columns,
            # so we loop over the header row to calculate the max column.
            max_column_idx = 0
            while sheet.cell(row=1, column=max_column_idx + 1).value is not None:
                max_column_idx += 1

            # Subtract one from column_idx since we will be iterating over a row, which is a 0-based
            # list, even though getting cells is a 1-based operation.
            while (cell := sheet.cell(row=1, column=column_idx)).value != 'PWFS1':
                instrument_column_mapping[cell.value] = column_idx - 1
                column_idx += 1

            while column_idx <= max_column_idx:
                wfs_columns[column_idx - 1] = sheet.cell(row=1, column=column_idx).value
                column_idx += 1

            # Skip the header row (row 1).
            for idxOffset, row in enumerate(sheet.iter_rows(min_row=2)):
                idx = idxOffset + 2

                # The error and logging messages all start with:
                msg = f'Configuration file for site {site} row {idx}'

                # Read the date and create an entry for the site and date.
                row_date = row[0].value.date()

                # Check the telescope status. If it is closed, we ignore the rest of the row.
                status = row[1].value.upper().strip()
                if status == 'CLOSED':
                    self._blocked[site].add(row_date)

                    continue
                elif status != 'OPEN':
                    raise ValueError(f'{msg} has illegal value in Telescope column: {status}.')

                # Process the mode.
                original_mode = row[2].value
                mode = original_mode.upper().strip()

                # We process the modes in the following order:
                # 1. Engineering / Shutdown (terminates)
                # 2. Visitor block <instrument-name>
                # 3. <partner> block
                # 4. PV: <prog-id-list>
                # 5. Classical: <prog-id-list>
                # 6. Priority: <prog-id-list>
                if mode in {FileBasedResourceService._ENGINEERING, 'SHUTDOWN'}:
                    self._blocked[site].add(row_date)
                    continue

                # Now we can add the date to dates since it will require further processing.
                dates.add(row_date)

                if mode.startswith('VISITOR:'):
                    instrument = self.lookup_resource(original_mode[8:].strip())
                    instrument_run.setdefault(instrument, set()).add(row_date)

                elif mode.startswith('PARTNER:'):
                    try:
                        partner = TimeAccountingCode[mode[8:].strip()]
                    except KeyError as ex:
                        raise KeyError(f'{msg} has illegal time account {ex} in mode: {mode}.')
                    partner_blocks[row_date] = partner

                elif mode.startswith('PV:'):
                    FileBasedResourceService._add_dates_to_dict(pv_programs, mode[3:], row_date)

                elif mode.startswith('CLASSICAL:'):
                    FileBasedResourceService._add_dates_to_dict(classical_programs, mode[10:], row_date)

                elif mode.startswith('PRIORITY:'):
                    FileBasedResourceService._add_dates_to_dict(score_boost, mode[9:], row_date)

                elif mode != 'QUEUE':
                    raise ValueError(f'{msg} has illegal mode: {mode}.')

                # Get the LGS status of the row. The default, if no LGS is specified, is False.
                lgs = str_to_bool(row[3].value)
                if not row[3].value:
                    logger.warning(f'{msg} has no LGS entry. Using default value of No.')
                    self._lgs[site][row_date] = False
                else:
                    self._lgs[site][row_date] = lgs

                # Get the ToO status for the night...
                too = str_to_bool(row[4].value)
                if not row[4].value:
                    logger.warning(f'{msg} has no ToO entry. Using default value of Yes.')
                    self._too[site][row_date] = True
                else:
                    self._too[site][row_date] = too

                # Determine instrument and WFS resources available.
                resources: Set[Resource] = set()

                # The site is available on the day, so add it to the resources.
                resources.add(site.resource)

                # The next five entries contain instrument ports.
                # Filter out any ports that are not empty.
                # We then have to check to see what the status of the instrument is for the night.
                instrument_names = {row[i].value for i in range(5, 10) if row[i].value is not None and row[i].value}

                for name in instrument_names:
                    if name not in instrument_column_mapping:
                        raise KeyError(f'{msg} contains illegal instrument name: {name}.')
                    try:
                        instrument_status = none_to_str(row[instrument_column_mapping[name]].value).strip().upper()
                    except IndexError:
                        # This happens if the row ends prematurely.
                        instrument_status = ''
                    if instrument_status == FileBasedResourceService._SCIENCE:
                        resources.add(self.lookup_resource(name))
                    elif not instrument_status:
                        logger.warning(f'{msg} contains no instrument status. Using default of Not Available.')
                    elif instrument_status not in [FileBasedResourceService._NOT_AVAILABLE,
                                                   FileBasedResourceService._ENGINEERING,
                                                   FileBasedResourceService._CALIBRATION]:
                        raise ValueError(f'{msg} for instrument {name} '
                                         f'contains illegal status: {instrument_status}.')

                # The remaining columns are WFS. Check the status and if available, add to resources.
                for idx, name in wfs_columns.items():
                    try:
                        wfs_status = none_to_str(row[idx].value).strip().upper()
                    except IndexError:
                        # This happens if the row ends prematurely.
                        wfs_status = ''
                    if wfs_status == FileBasedResourceService._SCIENCE:
                        resources.add(self.lookup_resource(name))
                    elif not wfs_status or wfs_status:
                        logger.warning(f'{msg} for WFS {name} contains no status. Using default of Not Available.')
                    elif wfs_status not in [FileBasedResourceService._NOT_AVAILABLE, FileBasedResourceService._ENGINEERING]:
                        raise ValueError(f'{msg} for WFS {name} contains illegal status: {wfs_status}.')

                # Add the resource data to the dates. Union returns a new set.
                self._resources[site][row_date] = self._resources[site].setdefault(row_date, set()).union(resources)

            # Determine the earliest and latest date for the site.
            self._earliest_date_per_site[site] = min(dates.union(self._blocked[site]))
            self._latest_date_per_site[site] = max(dates.union(self._blocked[site]))

            # Block out the blocked dates.
            for d in self._blocked[site]:
                s = self._positive_filters[site].setdefault(d, set())
                s.add(NothingFilter())

            # PV rules:
            for pv_program_id, pv_dates in pv_programs.items():
                pv_starting_date = min(pv_dates)

                # 1. PV programs get priority on the nights they are listed.
                for d in pv_dates:
                    s = self._positive_filters[site].setdefault(d, set())
                    s.add(ProgramPriorityFilter(frozenset({pv_program_id})))

                # 2. PV programs are not allowed in nights before they are first listed.
                pv_prohibited_dates = {d for d in dates if d < pv_starting_date}
                for d in pv_prohibited_dates:
                    s = self._negative_filters[site].setdefault(d, set())
                    s.add(ProgramPermissionFilter(frozenset({pv_program_id})))

            # Classical rules:
            for classical_program_id, classical_dates in classical_programs.items():
                for d in dates:
                    # Classical programs can only be performed in their designated blocks.
                    if d in classical_dates:
                        s = self._positive_filters[site].setdefault(d, set())
                        s.add(ProgramPriorityFilter(frozenset({classical_program_id})))
                    else:
                        s = self._negative_filters[site].setdefault(d, set())
                        s.add(ProgramPermissionFilter(frozenset({classical_program_id})))

            # Priority rules:
            for priority_program_id, priority_dates in score_boost.items():
                # Priority is given to programs in the priority block.
                for d in priority_dates:
                    s = self._positive_filters[site].setdefault(d, set())
                    s.add(ProgramPriorityFilter(frozenset({priority_program_id})))

            # Partner rules:
            for d, partner_code in partner_blocks.items():
                # On a partner night, we only allow programs that include the partner.
                s = self._positive_filters[site].setdefault(d, set())
                s.add(TimeAccountingCodeFilter(frozenset({partner_code})))

            # Visitor instrument rules:
            for resource, resource_dates in instrument_run.items():
                # Priority is given to scheduling blocks using the resource.
                for d in resource_dates:
                    s = self._positive_filters[site].setdefault(d, set())
                    s.add(ResourcePriorityFilter(frozenset({resource})))

            # ToO rules:
            for d in dates:
                # Block ToOs on nights where they are not allowed.
                if not self._too[site][d]:
                    s = self._negative_filters[site].setdefault(d, set())
                    s.add(TooFilter())

            # LGS rules:
            for d in dates:
                # Block LGS on nights where they are not allowed
                if not self._lgs[site][d]:
                    s = self._negative_filters[site].setdefault(d, set())
                    s.add(LgsFilter())

    @staticmethod
    def _mirror_parser(r: List[str], site: Site) -> Set[str]:
        return {'Mirror'} | {i.strip().replace('+', '') for i in r}

    def load_files(self,
                   site: Site,
                   fpu_to_barcodes_path: str,
                   fpus_path: Union[str, BytesIO],
                   gratings_path: Union[str, BytesIO],
                   spreadsheet: Union[str, BytesIO],
                   from_gdrive: bool = False):
        """
        Load all files necessaries to the correct functioning of the ResourceManager.
        """

        # Load the mappings from the ITCD FPU values to the barcodes.
        self._load_fpu_to_barcodes(site, fpu_to_barcodes_path)

        # Load the FPUrs.
        # This will put both the IFU and the FPU barcodes available on a given date as Resources.
        # Note that for the IFU, we need to convert to a barcode, which is a Resource.
        # This is a bit problematic since we expect a list of strings of Resource IDs, so we have to take its ID.
        self._load_csv(site,
                       self._itcd_fpu_to_barcode_parser,
                       fpus_path)

        # Load the gratings.
        # This will put the mirror and the grating names available on a given date as Resources.
        # TODO: Check Mirror vs. MIRROR. Seems like GMOS uses Mirror.
        self._load_csv(site,
                       self._mirror_parser,
                       gratings_path)

        # Process the spreadsheet information for instrument, mode, and LGS settings.
        self._load_spreadsheet(spreadsheet, from_gdrive=from_gdrive)
