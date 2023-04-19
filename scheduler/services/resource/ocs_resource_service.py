# Copyright (c) 2016-2022 Association of Universities for Research in Astronomy, Inc. (AURA)
# For license information see LICENSE or https://opensource.org/licenses/BSD-3-Clause

import csv
import re
import os
from copy import copy
from datetime import datetime, timedelta
from typing import Dict, Final, List, NoReturn, Set, Tuple

from astropy.time import Time
import gelidum
import requests
from lucupy.helpers import str_to_bool
from lucupy.minimodel import ALL_SITES
from lucupy.sky import night_events
from openpyxl import load_workbook

from definitions import ROOT_DIR
from scheduler.core.meta import Singleton
from scheduler.core.resourcemanager import ResourceManager
from scheduler.services import logger_factory
from .filters import *
from .google_drive_downloader import GoogleDriveDownloader
from .night_resource_configuration import *


logger = logger_factory.create_logger(__name__)


@final
class OcsResourceService(ResourceManager, metaclass=Singleton):
    """
    This is a mock for the future Resource service, used for OCS.
    It reads data regarding availability of instruments, IFUs, FPUs, MOS masks, etc. at each Site for given dates.

    It can then be queried to receive a set of Resource (usually with barcode IDs, except for instruments) for a
    given site on a given night.

    It caches and reuses Resources by ID as best as possible to minimize the number of Resource objects existing
    at any given time. Since Resource is immutable, this should be fine.

    Note that this is a Singleton class, so new instances do not need to be created.
    """
    # The Google ID of the telescope configuration file.
    _SITE_CONFIG_GOOGLE_ID: Final[str] = '1QRalQNEaX-bcyrPG6mfKnv01JVMaGHwy'

    # Name of the spreadsheet file containing telescope configurations.
    _SITE_CONFIG_FILE: Final[str] = 'telescope_schedules.xlsx'

    # Definition of a day to not have to redeclare constantly.
    _day: Final[timedelta] = timedelta(days=1)

    # Statuses of instruments and WFS.
    _SCIENCE: Final[str] = 'SCIENCE'
    _ENGINEERING: Final[str] = 'ENGINEERING'
    _NOT_AVAILABLE: Final[str] = 'NOT AVAILABLE'
    _CALIBRATION: Final[str] = 'CALIBRATION'

    # These are the converters from the OCS FPU names to the ITCD FPU representations.
    # For example, the ODB query extractor would return:
    #    * 'IFU 2 Slits'
    # which we would want to convert to:
    #    * 'IFU-2'
    # since these are the FPU names used in the GMOS[NS]-FPUr######.txt files.
    _gmosn_ifu_dict: Dict[str, str] = gelidum.freeze({
        'IFU 2 Slits': 'IFU-2',
        'IFU Left Slit (blue)': 'IFU-B',
        'IFU Right Slit (red)': 'IFU-R',
        'Longslit 0.25 arcsec': '0.25arcsec',
        'Longslit 0.50 arcsec': '0.5arcsec',
        'Longslit 0.75 arcsec': '0.75arcsec',
        'Longslit 1.00 arcsec': '1.0arcsec',
        'Longslit 1.5 arcsec': '1.5arcsec',
        'Longslit 2.00 arcsec': '2.0arcsec',
        'Longslit 5.00 arcsec': '5.0arcsec',
        'N and S 0.50 arcsec': 'NS0.5arcsec',
        'N and S 0.75 arcsec': 'NS0.75arcsec',
        'N and S 1.00 arcsec': 'NS1.0arcsec',
        'N and S 1.50 arcsec': 'NS1.5arcsec',
        'N and S 2.00 arcsec': 'NS2.0arcsec',
        'focus_array_new': 'focus_array_new'
    })

    _gmoss_ifu_dict: Dict[str, str] = gelidum.freeze({**_gmosn_ifu_dict, **{
        'IFU N and S 2 Slits': 'IFU-NS-2',
        'IFU N and S Left Slit (blue)': 'IFU-NS-B',
        'IFU N and S Right Slit (red)': 'IFU-NS-R',
        # TODO: Not in OCS?
        'PinholeC': 'PinholeC'
    }})

    def __init__(self, sites: FrozenSet[Site] = ALL_SITES):
        """
        Create and initialize the ResourceMock object with the specified sites.
        """
        super().__init__()
        self._sites = sites
        self._path = os.path.join(ROOT_DIR, 'scheduler', 'services', 'resource', 'data')

        # The map from site and date to the set of resources.
        self._resources: Dict[Site, Dict[date, Set[Resource]]] = {site: {} for site in self._sites}

        # Mapping from ITCD FPUs to barcodes. The mapping is site-dependent.
        # The ODB program extractor produces long versions of these names that must be run through the
        # OcsFpuConverter to get the ITCD FPU names.
        self._itcd_fpu_to_barcode: Dict[Site, Dict[str, Resource]] = {site: {} for site in self._sites}

        # Earliest and latest dates per site.
        self._earliest_date_per_site: Dict[Site, date] = {site: date.max for site in self._sites}
        self._latest_date_per_site: Dict[Site, date] = {site: date.min for site in self._sites}

        # Determines which nights are blocked.
        self._blocked: Dict[Site, Set[date]] = {site: set() for site in self._sites}

        # Determines whether a night is a part of a laser run.
        self._lgs: Dict[Site, Dict[date, bool]] = {site: {} for site in self._sites}

        # Determines whether ToOs are accepted on a given night.
        self._too: Dict[Site, Dict[date, bool]] = {site: {} for site in self._sites}

        # Fault reports by datetime to calculate missing instruments
        self._faults: Dict[Site, Dict[date, Fault]] = {site: {} for site in self._sites}

        # Engineering Task by datetime.
        self._eng_task: Dict[Site, Dict[date, EngTask]] = {site: {} for site in self._sites}

        # Filters to apply to a night. We add the ResourceFilters at the end after all the resources
        # have been processed.
        self._positive_filters: Dict[Site, Dict[date, Set[AbstractFilter]]] = {site: {} for site in self._sites}
        self._negative_filters: Dict[Site, Dict[date, Set[AbstractFilter]]] = {site: {} for site in self._sites}

        # The final combined filters for each night.
        self._filters: Dict[Site, Dict[date, AbstractFilter]] = {site: {} for site in self._sites}

        # The final output from this class: the configuration per night.
        self._night_configurations: Dict[Site, Dict[date, NightConfiguration]] = {site: {} for site in self._sites}

        for site in self._sites:
            suffix = ('s' if site == Site.GS else 'n').upper()

            # Load the mappings from the ITCD FPU values to the barcodes.
            self._load_fpu_to_barcodes(site, f'GMOS{suffix}_fpu_barcode.txt')

            # Load the FPUrs.
            # This will put both the IFU and the FPU barcodes available on a given date as Resources.
            # Note that for the IFU, we need to convert to a barcode, which is a Resource.
            # This is a bit problematic since we expect a list of strings of Resource IDs, so we have to take its ID.
            self._load_csv(site, f'GMOS{suffix}_FPUr201789.txt',
                           lambda r: {self._itcd_fpu_to_barcode[site][r[0].strip()].id} | {i.strip() for i in r[1:]})

            # Load the gratings.
            # This will put the mirror and the grating names available on a given date as Resources.
            # TODO: Check Mirror vs. MIRROR. Seems like GMOS uses Mirror.
            self._load_csv(site, f'GMOS{suffix}_GRAT201789.txt',
                           lambda r: {'Mirror'} | {i.strip().replace('+', '') for i in r})

            # Load faults files.
            self.parse_faults_file(site, f'Faults_AllG{suffix}.txt')

            # Load eng task files.
            self._parse_eng_task_file(site, f'ClosedDomeG{suffix}_Eng.txt')

        # Process the spreadsheet information for instrument, mode, and LGS settings.
        self._load_spreadsheet()

        # TODO: Remove this after discussion with science.
        # TODO: There are entries here outside of the Telescope Schedules Spreadsheet.
        # Record the earliest date for each site: any date before this will return an empty set of Resources.
        # Record the latest date for each site: any date after this will return the Resources on this date.
        # self._earliest_date_per_site = {site: min(self._resources[site], default=None) for site in self._sites}
        # self._latest_date_per_site = {site: max(self._resources[site], default=None) for site in self._sites}
        for site in self._sites:
            # Only one of these checks should be necessary.
            if self._earliest_date_per_site[site] == date.max or self._latest_date_per_site[site] == date.min:
                raise ValueError(f'No site resource data for {site.name}.')

        # Finalize the filters and create the night configurations.
        for site in self._sites:
            d = self._earliest_date_per_site[site]
            while d <= self._latest_date_per_site[site]:
                # Now that we have a complete set of resources per night:
                # 1. Make sure that there are entries in the positive_filters and negative_filters for the date.
                # 2. Add the ResourceFilter to the positive filters.
                # 2. Combine into a composite filter.
                pf = self._positive_filters[site].setdefault(d, set())
                nf = self._negative_filters[site].setdefault(d, set())
                pf.add(ResourcesAvailableFilter(frozenset(self._resources[site][d])))
                composite_filter = CompositeFilter(frozenset(pf), frozenset(nf))

                self._night_configurations[site][d] = NightConfiguration(
                    site=site,
                    local_date=d,
                    is_lgs=(d not in self._blocked[site] and self._lgs[site][d]),
                    too_status=(d not in self._blocked[site] and self._too[site][d]),
                    filter=composite_filter,
                    resources=frozenset(self._resources[site][d]),
                    faults=frozenset(self._faults[site][d]) if d in self._faults[site] else frozenset(),
                    eng_tasks=frozenset(self._faults[site][d]) if d in self._faults[site] else frozenset()
                )

                d += OcsResourceService._day

    def _load_fpu_to_barcodes(self, site: Site, name: str) -> NoReturn:
        """
        FPUs at each site map to a unique barcode as defined in the files:
            * gmos[ns]_fpu_barcode.txt
        These are site-dependent values.
        """
        with open(os.path.join(self._path, name)) as f:
            for row in f:
                fpu, barcode = row.split()

                # Only map if the FPU is a resource.
                if fpu is not None:
                    self._itcd_fpu_to_barcode[site][fpu] = self.lookup_resource(barcode)

    def _load_csv(self, site: Site, name: str, c: Callable[[List[str]], Set[str]]) -> NoReturn:
        """
        Process a CSV file as a table, where:

        1. The first entry is a date in YYYY-mm-dd format
        2. The remaining entries are resources available on that date to the following date in the file.

        If a date is missing from the CSV file, copy the data from the previously defined date through to just before
        the new date.
        """
        with open(os.path.join(self._path, name)) as f:
            reader = csv.reader(f, delimiter=',')
            prev_row_date: Optional[date] = None

            for row in reader:
                row_date = datetime.strptime(row[0].strip(), '%Y-%m-%d').date()

                # Fill in any gaps by copying prev_row_date until we reach one less than row_date.
                if prev_row_date is not None:
                    missing_row_date = prev_row_date + OcsResourceService._day
                    while missing_row_date < row_date:
                        # Make sure there is an entry and append to it to avoid overwriting anything already present.
                        date_set = self._resources[site].setdefault(missing_row_date, set())
                        self._resources[site][missing_row_date] = date_set | copy(self._resources[site][prev_row_date])
                        missing_row_date += OcsResourceService._day

                # Get or create date_set for the date, and append new resources from table, ignoring blank entries.
                date_set = self._resources[site].setdefault(row_date, set())
                new_entries = {self.lookup_resource(r) for r in c(row[1:]) if r}
                self._resources[site][row_date] = date_set | new_entries

                # Advance the previous row date where data was defined.
                prev_row_date = row_date

    @staticmethod
    def _split_program_ids(text: str) -> Set[ProgramID]:
        """
        Given a text list that should contain comma-separated program IDs, separate them
        and return the set of them.
        """
        return set(prog_id.strip() for prog_id in text.strip().split(','))

    @staticmethod
    def _add_dates_to_dict(prog_dict: Dict[ProgramID, Set[date]], text: str, add_date: date) -> NoReturn:
        """
        Given a program dictionary, parse the text following it into program IDs and add the
        specified date to each set associated with each program ID.
        """
        for prog_id in OcsResourceService._split_program_ids(text):
            date_set = prog_dict.setdefault(prog_id, set())
            date_set.add(add_date)

    def _load_spreadsheet(self) -> NoReturn:
        """
        Process an Excel spreadsheet containing instrument, mode, and LGS information.

        The Excel spreadsheets have information available for every date, so we do not have to concern ourselves
        as in the _load_csv file above.
        """
        def none_to_str(value) -> str:
            return '' if value is None else value

        filename = os.path.join(self._path, OcsResourceService._SITE_CONFIG_FILE)

        logger.info('Retrieving site configuration file from Google Drive...')
        try:
            GoogleDriveDownloader.download_file_from_google_drive(file_id=OcsResourceService._SITE_CONFIG_GOOGLE_ID,
                                                                  overwrite=True,
                                                                  dest_path=filename)
        except requests.RequestException:
            logger.warning('Could not retrieve site configuration file from Google Drive.')

        if not os.path.exists(filename):
            raise FileNotFoundError(f'No site configuration data available for {__class__.__name__} at: {filename}')

        workbook = load_workbook(filename=filename, read_only=True, data_only=True)
        for site in self._sites:
            try:
                sheet = workbook[site.name]
            except KeyError:
                # Make the KeyError more clear.
                raise KeyError(f'{__class__.__name__}: No tab for {site.name} in the '
                               'telescope schedule configuration spreadsheet.')

            # A set consisting of the dates that are not blocked.
            dates: Set[date] = set()

            # We need to determine what programs are prohibited from what nights.
            # PV programs can only be done on PV nights or later, so we store the night they first appear.
            # They also receive boosted pri
            pv_programs: Dict[ProgramID, Set[date]] = {}

            # Classical programs can only be done on specifically designated nights.
            classical_programs: Dict[ProgramID, Set[date]] = {}

            # Programs that get a boost on dates.
            score_boost: Dict[ProgramID, Set[date]] = {}

            # Partner blocks are when the mode indicates that the night is restricted to a partner.
            partner_blocks: Dict[date, TimeAccountingCode] = {}

            # Instrument runs, where observations associated with certain instruments get a boost.
            instrument_run: Dict[Resource, Set[date]] = {}

            # Process the header.
            # We need to make mappings from all the instrument names to their column number.
            # We also need the names of the WFS columns. These are reversed because we want to go from column idx to
            #   resource.
            instrument_column_mapping: Dict[str, int] = {}
            wfs_columns: Dict[int, str] = {}

            column_idx = 11

            # sheet.max_column is returning None, probably because of missing data in PWFS columns,
            # so we loop over the header row to calculate the max column.
            max_column_idx = 0
            while sheet.cell(row=1, column=max_column_idx+1).value is not None:
                max_column_idx += 1

            # Subtract one from column_idx since we will be iterating over a row, which is a 0-based
            # list, even though getting cells is a 1-based operation.
            while (cell := sheet.cell(row=1, column=column_idx)).value != 'PWFS1':
                instrument_column_mapping[cell.value] = column_idx - 1
                column_idx += 1

            while column_idx <= max_column_idx:
                wfs_columns[column_idx - 1] = sheet.cell(row=1, column=column_idx).value
                column_idx += 1

            # Skip the header row (row 1).
            for idxOffset, row in enumerate(sheet.iter_rows(min_row=2)):
                idx = idxOffset + 2

                # The error and logging messages all start with:
                msg = f'Configuration file for site {site} row {idx}'

                # Read the date and create an entry for the site and date.
                row_date = row[0].value.date()

                # Check the telescope status. If it is closed, we ignore the rest of the row.
                status = row[1].value.upper().strip()
                if status == 'CLOSED':
                    self._blocked[site].add(row_date)

                    continue
                elif status != 'OPEN':
                    raise ValueError(f'{msg} has illegal value in Telescope column: {status}.')

                # Process the mode.
                original_mode = row[2].value
                mode = original_mode.upper().strip()

                # We process the modes in the following order:
                # 1. Engineering / Shutdown (terminates)
                # 2. Visitor block <instrument-name>
                # 3. <partner> block
                # 4. PV: <prog-id-list>
                # 5. Classical: <prog-id-list>
                # 6. Priority: <prog-id-list>
                if mode in {OcsResourceService._ENGINEERING, 'SHUTDOWN'}:
                    self._blocked[site].add(row_date)
                    continue

                # Now we can add the date to dates since it will require further processing.
                dates.add(row_date)

                if mode.startswith('VISITOR:'):
                    instrument = self.lookup_resource(original_mode[8:].strip())
                    instrument_run.setdefault(instrument, set()).add(row_date)

                elif mode.startswith('PARTNER:'):
                    try:
                        partner = TimeAccountingCode[mode[8:].strip()]
                    except KeyError as ex:
                        raise KeyError(f'{msg} has illegal time account {ex} in mode: {mode}.')
                    partner_blocks[row_date] = partner

                elif mode.startswith('PV:'):
                    OcsResourceService._add_dates_to_dict(pv_programs, mode[3:], row_date)

                elif mode.startswith('CLASSICAL:'):
                    OcsResourceService._add_dates_to_dict(classical_programs, mode[10:], row_date)

                elif mode.startswith('PRIORITY:'):
                    OcsResourceService._add_dates_to_dict(score_boost, mode[9:], row_date)

                elif mode != 'QUEUE':
                    raise ValueError(f'{msg} has illegal mode: {mode}.')

                # Get the LGS status of the row. The default, if no LGS is specified, is False.
                lgs = str_to_bool(row[3].value)
                if not row[3].value:
                    logger.warning(f'{msg} has no LGS entry. Using default value of No.')
                    self._lgs[site][row_date] = False
                else:
                    self._lgs[site][row_date] = lgs

                # Get the ToO status for the night..
                too = str_to_bool(row[4].value)
                if not row[4].value:
                    logger.warning(f'{msg} has no ToO entry. Using default value of Yes.')
                    self._too[site][row_date] = True
                else:
                    self._too[site][row_date] = too

                # Determine instrument and WFS resources available.
                resources: Set[Resource] = set()

                # The next five entries contain instrument ports.
                # Filter out any ports that are not empty.
                # We then have to check to see what the status of the instrument is for the night.
                instrument_names = {row[i].value for i in range(5, 10) if row[i].value is not None and row[i].value}

                for name in instrument_names:
                    if name not in instrument_column_mapping:
                        raise KeyError(f'{msg} contains illegal instrument name: {name}.')
                    try:
                        instrument_status = none_to_str(row[instrument_column_mapping[name]].value).strip().upper()
                    except IndexError:
                        # This happens if the row ends prematurely.
                        instrument_status = ''
                    if instrument_status == OcsResourceService._SCIENCE:
                        resources.add(self.lookup_resource(name))
                    elif not instrument_status:
                        logger.warning(f'{msg} contains no instrument status. Using default of Not Available.')
                    elif instrument_status not in [OcsResourceService._NOT_AVAILABLE,
                                                   OcsResourceService._ENGINEERING,
                                                   OcsResourceService._CALIBRATION]:
                        raise ValueError(f'{msg} for instrument {name} '
                                         f'contains illegal status: {instrument_status}.')

                # The remaining columns are WFS. Check the status and if available, add to resources.
                for idx, name in wfs_columns.items():
                    try:
                        wfs_status = none_to_str(row[idx].value).strip().upper()
                    except IndexError:
                        # This happens if the row ends prematurely.
                        wfs_status = ''
                    if wfs_status == OcsResourceService._SCIENCE:
                        resources.add(self.lookup_resource(name))
                    elif not wfs_status or wfs_status:
                        logger.warning(f'{msg} for WFS {name} contains no status. Using default of Not Available.')
                    elif wfs_status not in [OcsResourceService._NOT_AVAILABLE, OcsResourceService._ENGINEERING]:
                        raise ValueError(f'{msg} for WFS {name} contains illegal status: {wfs_status}.')

                # Add the resource data to the dates. Union returns a new set.
                self._resources[site][row_date] = self._resources[site].setdefault(row_date, set()).union(resources)

            # Determine the earliest and latest date for the site.
            self._earliest_date_per_site[site] = min(dates.union(self._blocked[site]))
            self._latest_date_per_site[site] = max(dates.union(self._blocked[site]))

            # Block out the blocked dates.
            for d in self._blocked[site]:
                s = self._positive_filters[site].setdefault(d, set())
                s.add(NothingFilter())

            # PV rules:
            for pv_program_id, pv_dates in pv_programs.items():
                pv_starting_date = min(pv_dates)

                # 1. PV programs get priority on the nights they are listed.
                for d in pv_dates:
                    s = self._positive_filters[site].setdefault(d, set())
                    s.add(ProgramPriorityFilter(frozenset(pv_program_id)))

                # 2. PV programs are not allowed in nights before they are first listed.
                pv_prohibited_dates = {d for d in dates if d < pv_starting_date}
                for d in pv_prohibited_dates:
                    s = self._negative_filters[site].setdefault(d, set())
                    s.add(ProgramPermissionFilter(frozenset([pv_program_id])))

            # Classical rules:
            for classical_program_id, classical_dates in classical_programs.items():
                for d in dates:
                    # Classical programs can only be performed in their designated blocks.
                    if d in classical_dates:
                        s = self._positive_filters[site].setdefault(d, set())
                        s.add(ProgramPriorityFilter(frozenset([classical_program_id])))
                    else:
                        s = self._negative_filters[site].setdefault(d, set())
                        s.add(ProgramPermissionFilter(frozenset([classical_program_id])))

            # Priority rules:
            for priority_program_id, priority_dates in score_boost.items():
                # Priority is given to programs in the priority block.
                for d in priority_dates:
                    s = self._positive_filters[site].setdefault(d, set())
                    s.add(ProgramPriorityFilter(frozenset(priority_program_id)))

            # Partner rules:
            for d, partner_code in partner_blocks.items():
                # On a partner night, we only allow programs that include the partner.
                s = self._positive_filters[site].setdefault(d, set())
                s.add(TimeAccountingCodeFilter(frozenset(partner_code)))

            # Visitor instrument rules:
            for resource, resource_dates in instrument_run.items():
                # Priority is given to scheduling blocks using the resource.
                for d in resource_dates:
                    s = self._positive_filters[site].setdefault(d, set())
                    s.add(ResourcePriorityFilter(frozenset({resource})))

            # ToO rules:
            for d in dates:
                # Block ToOs on nights where they are not allowed.
                if not self._too[site][d]:
                    s = self._negative_filters[site].setdefault(d, set())
                    s.add(TooFilter())

            # LGS rules:
            for d in dates:
                # Block LGS on nights where they are not allowed
                if not self._lgs[site][d]:
                    s = self._negative_filters[site].setdefault(d, set())
                    s.add(LgsFilter())

    def _parse_eng_task_file(self, site: Site, to_file: str) -> None:
        """
        Parse Engineering task that block moments or the entire night.
        Each twilight is calculated using lucupy.sky, some discrepancies might affect.
        """
        pattern = r'(\[.*\])'
        # Ignore GS until the file is created.
        if site is Site.GS:
            return

        with open(os.path.join(self._path, to_file), 'r') as file:
            for line in file:
                # Find pattern to keep bracket comments not splitted
                match = re.search(pattern, line)
                if match:
                    comment = match.group(1)
                    rest_of_line = re.sub(pattern, '', line)
                    _date, start_time, end_time = rest_of_line.strip().split()
                    #  Single date for key
                    just_date = datetime.strptime(_date, '%Y-%m-%d')
                    # Time day in jd to calculate twilights
                    time = Time(just_date)
                    _, _, _, even_12twi, morn_12twi, _, _ = night_events(time,
                                                                         site.location,
                                                                         site.timezone)

                    # Handle twilights
                    if start_time == 'twi':
                        _start_date = even_12twi.datetime
                        _end_date = morn_12twi.datetime if end_time == 'twi' else datetime.strptime(f'{_date} {end_time}',
                                                                                     '%Y-%m-%d %H:%M')
                    else:
                        _start_date = datetime.strptime(f'{_date} {start_time}',
                                                        '%Y-%m-%d %H:%M')
                        _end_date = morn_12twi.datetime if end_time == 'twi' else datetime.strptime(f'{_date} {end_time}',
                                                                                     '%Y-%m-%d %H:%M')

                    time_loss = _end_date - _start_date
                    eng_task = EngTask(start=_start_date,
                                       end=_end_date,
                                       time_loss=time_loss,
                                       comment=comment)

                    if just_date in self._eng_task[site]:
                        self._eng_task[site][just_date].append(eng_task)
                    else:
                        self._eng_task[site][just_date] = [eng_task]
                else:
                    raise ValueError('Pattern not found. Format error on Eng Task file')

    def parse_faults_file(self, site: Site, to_file: str)-> None:
        """Parse faults from files.
        This is purposeful left non-private as might be used with incoming files from
        the React app.
        """
        # Files contains this repetitive string in each timestamp, if we need them
        # we could add them as constants.
        ts_clean = ' 04:00' if site == Site.GS else ' 10:00'
        with open(os.path.join(self._path, to_file), 'r') as file:
            for l in file:
                line = l.rstrip()  # remove trail spaces
                if line:  # ignore empty lines
                    if line[0].isdigit():
                        semester = line
                    elif line.startswith('FR'):  # found a fault
                        items = line.split('\t')
                        # Create timestamp with ts_clean var removed
                        ts = datetime.strptime(items[1].replace(ts_clean, ''),
                                               '%Y %m %d  %H:%M:%S')
                        fault = Fault(items[0],
                                      ts,  # date with time
                                      timedelta(hours=float(items[2])),  # timeloss
                                      items[3]) # comment for the fault
                        if ts.date() in self._faults[site]:
                            self._faults[site][ts.date()].append(fault)
                        else:
                            self._faults[site][ts.date()] = [fault]
                    else:
                        raise ValueError('Fault file has wrong format')

    def date_range_for_site(self, site: Site) -> Tuple[date, date]:
        """
        Return the date range (inclusive) for which we have resource data for a site.
        """
        if site not in self._sites:
            raise ValueError(f'Request for resource dates for illegal site: {site.name}')
        return self._earliest_date_per_site[site], self._latest_date_per_site[site]

    def get_night_configuration(self, site: Site, local_date: date) -> NightConfiguration:
        """
        Returns the NightConfiguration object for the site for the given local date,
        which contains the filters and resources for the night.
        """
        if site not in self._sites:
            raise ValueError(f'Request for night configuration for illegal site: {site.name}')
        if local_date < self._earliest_date_per_site[site] or local_date > self._latest_date_per_site[site]:
            raise ValueError(f'Request for night configuration for site {site.name} for illeegal date: {local_date}')
        return self._night_configurations[site][local_date]

    def get_resources(self, site: Site, night_date: date) -> FrozenSet[Resource]:
        """
        For a site and a local date, return the set of available resources.
        The date is currently the truncation to day of the astropy Time objects in the time_grid, which are in UTC
           and have times of 8:00 am, so to get local dates, in the Collector we subtract one day.
        If the date falls before any resource data for the site, return the empty set.
        If the date falls after any resource data for the site, return the last resource set.
        """
        if site not in self._sites:
            raise ValueError(f'Request for resources for illegal site: {site.name}')

        # If the date is before the first date or after the last date, return the empty set.
        if night_date < self._earliest_date_per_site[site] or night_date > self._latest_date_per_site[site]:
            return frozenset()

        return frozenset(self._resources[site][night_date])

    def fpu_to_barcode(self, site: Site, fpu_name: str) -> Optional[Resource]:
        """
        Convert a long FPU name into the barcode, if it exists.
        """
        itcd_fpu_name = self._convert_fpu_to_itcd_name(site, fpu_name)
        return self._itcd_fpu_to_barcode[site].get(itcd_fpu_name)

    def _convert_fpu_to_itcd_name(self, site: Site, fpu_name: str) -> Optional[str]:
        """
        Convert a long FPU name into its ITCD name, if it exists.
        """
        if Site.GN in self._sites and site == Site.GN:
            return self._gmosn_ifu_dict.get(fpu_name)
        if Site.GS in self._sites and site == Site.GS:
            return self._gmoss_ifu_dict.get(fpu_name)
        return None
