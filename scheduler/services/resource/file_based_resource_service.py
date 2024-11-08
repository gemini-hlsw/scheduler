# Copyright (c) 2016-2024 Association of Universities for Research in Astronomy, Inc. (AURA)
# For license information see LICENSE or https://opensource.org/licenses/BSD-3-Clause

import csv
import re
from copy import copy
from datetime import date, datetime, time, timedelta
from io import BytesIO, StringIO
from typing import Callable, Dict, Final, List, Optional, Set, Type, TypeVar, Union

from astropy.time import Time
from openpyxl.reader.excel import load_workbook
from lucupy.helpers import str_to_bool
from lucupy.minimodel import ProgramID, Resource, Site, TimeAccountingCode, ResourceType
from lucupy.sky import night_events

from scheduler.services import logger_factory
from .event_generators import EngineeringTask, Fault, WeatherClosure
from .filters import (LgsFilter, NothingFilter, ProgramPermissionFilter, ProgramPriorityFilter, ResourcePriorityFilter,
                      TimeAccountingCodeFilter, TooFilter)
from .resource_service import ResourceService


__all__ = [
    'FileBasedResourceService',
]

T = TypeVar('T', EngineeringTask, Fault, WeatherClosure)

logger = logger_factory.create_logger(__name__)


class FileBasedResourceService(ResourceService):
    """
    This is the base for both OCS and File-uploaded services
    """
    # Definition of a day to not have to redeclare constantly.
    _day: Final[timedelta] = timedelta(days=1)
    _noon: Final[time] = time(hour=12, minute=0)

    # Statuses of instruments and WFS.
    _SCIENCE: Final[str] = 'SCIENCE'
    _ENGINEERING: Final[str] = 'ENGINEERING'
    _SHUTDOWN: Final[str] = 'SHUTDOWN'
    _NOT_AVAILABLE: Final[str] = 'NOT AVAILABLE'
    _CALIBRATION: Final[str] = 'CALIBRATION'

    def _load_fpu_to_barcodes(self, site: Site,
                              filename: str) -> None:
        """
        FPUs at each site map to a unique barcode as defined in the files:
            * gmos[ns]_fpu_barcode.txt
        These are site-dependent values.
        """
        with open(self._common / filename) as f:
            for row in f:
                fpu, barcode = row.split()

                # Only map if the FPU is a resource.
                if fpu is not None:
                    self._itcd_fpu_to_barcode[site][fpu] = self.lookup_resource(barcode,
                                                                                description=fpu,
                                                                                resource_type=ResourceType.FPU)

    def _load_csv(self,
                  site: Site,
                  c: Callable[[List[str], Site], Set[str]],
                  data_source: Union[str, BytesIO],
                  desc: Optional[str] = None,
                  resource_type: Optional[ResourceType] = ResourceType.NONE) -> None:
        """
        Process a CSV file as a table, where:

        1. The first entry is a date in YYYY-mm-dd format
        2. The remaining entries are resources available on that date to the following date in the file.

        If a date is missing from the CSV file, copy the data from the previously defined date through to just before
        the new date.
        """
        def _process_file(io) -> None:
            reader = csv.reader(io, delimiter=',')
            prev_row_date: Optional[date] = None

            for row in reader:
                # Get rid of the byte-order marker, which causes datetime.strptime to fail.
                row = [col.replace('\ufeff', '') for col in row]
                row_date = datetime.strptime(row[0].strip(), '%Y-%m-%d').date()

                # Fill in any gaps by copying prev_row_date until we reach one less than row_date.
                if prev_row_date is not None:
                    missing_row_date = prev_row_date + FileBasedResourceService._day
                    while missing_row_date < row_date:
                        # Make sure there is an entry and append to it to avoid overwriting anything already present.
                        date_set = self._resources[site].setdefault(missing_row_date, set())
                        self._resources[site][missing_row_date] = date_set | copy(self._resources[site][prev_row_date])
                        missing_row_date += FileBasedResourceService._day

                # Get or create date_set for the date, and append new resources from table, ignoring blank entries.
                date_set = self._resources[site].setdefault(row_date, set())
                # new_entries = {self.lookup_resource(r) for r in c(row[1:], site) if r}
                new_entries = {self.lookup_resource(r, description=desc, resource_type=resource_type)
                               for r in c(row[1:], site) if r}
                self._resources[site][row_date] = date_set | new_entries

                # Advance the previous row date where data was defined.
                prev_row_date = row_date

        if isinstance(data_source, str):
            with open(self._subdir / data_source) as f:
                _process_file(f)
        else:
            data_str = data_source.getvalue().decode()
            f = StringIO(data_str)
            _process_file(f)

    @staticmethod
    def _split_program_ids(text: str) -> Set[ProgramID]:
        """
        Given a text list that should contain comma-separated program IDs, separate them
        and return the set of them.
        """
        return set(ProgramID(prog_id.strip()) for prog_id in text.strip().split(','))

    @staticmethod
    def _add_dates_to_dict(prog_dict: Dict[ProgramID, Set[date]], text: str, add_date: date) -> None:
        """
        Given a program dictionary, parse the text following it into program IDs and add the
        specified date to each set associated with each program ID.
        """
        for prog_id in FileBasedResourceService._split_program_ids(text):
            date_set = prog_dict.setdefault(prog_id, set())
            date_set.add(add_date)

    @staticmethod
    def _add_dates_to_partdict(part_dict: Dict[TimeAccountingCode, Set[date]],
                               code: TimeAccountingCode, add_date: date) -> None:
        """
        Add the specified date to each set associated with the partner code.
        """
        date_set = part_dict.setdefault(code, set())
        date_set.add(add_date)

    def _load_instrument_data(self,
                              site: Site,
                              filename: str) -> None:
        """
        Process an Excel spreadsheet containing instrument, mode, and LGS information.

        The Excel spreadsheets have information available for every date, so we do not have to concern ourselves
        as in the _load_csv file above.
        """
        def none_to_str(value) -> str:
            return '' if value is None else value

        if not filename:
            raise ValueError('file_source cannot be empty')

        file_path = self._subdir / filename
        if not file_path.exists():
            raise FileNotFoundError(f'No site configuration data available for {__class__.__name__} at: '
                                    f'{file_path}')

        workbook = load_workbook(filename=file_path,
                                 read_only=True,
                                 data_only=True)

        try:
            sheet = workbook[site.name]
        except KeyError:
            # Make the KeyError more clear.
            raise KeyError(f'{__class__.__name__}: No tab for {site.name} in the '
                           'telescope schedule configuration spreadsheet.')

        # A set consisting of the dates that are not blocked.
        dates: Set[date] = set()

        # We need to determine what programs are prohibited from what nights.
        # PV programs can only be done on PV nights or later, so we store the night they first appear.
        # They also receive boosted pri
        pv_programs: Dict[ProgramID, Set[date]] = {}

        # Classical programs can only be done on specifically designated nights.
        classical_programs: Dict[ProgramID, Set[date]] = {}

        # Programs that get a boost on dates.
        score_boost: Dict[ProgramID, Set[date]] = {}

        # Partner blocks are when the mode indicates that the night is restricted to a partner.
        partner_blocks: Dict[TimeAccountingCode, Set[date]] = {}

        # Instrument runs, where observations associated with certain instruments get a boost.
        instrument_run: Dict[Resource, Set[date]] = {}

        # Process the header.
        # We need to make mappings from all the instrument names to their column number.
        # We also need the names of the WFS columns. These are reversed because we want to go from column idx to
        #   resource.
        instrument_column_mapping: Dict[str, int] = {}
        wfs_columns: Dict[int, str] = {}

        column_idx = 11

        # sheet.max_column is returning None, probably because of missing data in PWFS columns,
        # so we loop over the header row to calculate the max column.
        max_column_idx = 0
        while sheet.cell(row=1, column=max_column_idx + 1).value is not None:
            max_column_idx += 1

        # Subtract one from column_idx since we will be iterating over a row, which is a 0-based
        # list, even though getting cells is a 1-based operation.
        while (cell := sheet.cell(row=1, column=column_idx)).value != 'PWFS1':
            instrument_column_mapping[cell.value] = column_idx - 1
            column_idx += 1

        while column_idx <= max_column_idx:
            wfs_columns[column_idx - 1] = sheet.cell(row=1, column=column_idx).value
            column_idx += 1

        # Skip the header row (row 1).
        for idxOffset, row in enumerate(sheet.iter_rows(min_row=2)):
            idx = idxOffset + 2

            # The error and logging messages all start with:
            msg = f'Configuration file for site {site} row {idx}'

            # Read the date and create an entry for the site and date.
            row_date = row[0].value.date()
            # print(f'{row_date}')

            # Check the telescope status. If it is closed, we ignore the rest of the row.
            status = row[1].value.upper().strip()
            if status == 'CLOSED':
                self._blocked[site].add(row_date)
                continue

            elif status != 'OPEN':
                raise ValueError(f'{msg} has illegal value in Telescope column: {status}.')

            # Process the mode entries. There may be more than one, separated by |.
            modes_entry = row[2].value

            # Remove all parenthetical expressions from modes_entry.
            # TODO: We may have to handle these at some point.
            # Substitute the parenthetical expressions with an empty string.
            # We handle all in uppercase to reduce need for absolute precision.
            # Partner codes, program IDs, and instrument names are all in uppercase anyway.
            modes_entry = re.sub(r"\([^)]*\)", "", modes_entry).upper().strip()

            # If ENGINEERING or SHUTDOWN is in the mode entries, there is nothing else to do for the night.
            if (FileBasedResourceService._ENGINEERING in modes_entry
                    or FileBasedResourceService._SHUTDOWN in modes_entry):
                self._blocked[site].add(row_date)
                continue

            # Now we can add the date to dates since it will require further processing.
            dates.add(row_date)

            for mode_entry in (entry.strip() for entry in modes_entry.split('|')):
                # We process the modes in the following order:
                # 1. Visitor block <instrument-name>
                # 2. <partner> block
                # 3. PV: <prog-id-list>
                # 4. Classical: <prog-id-list>
                # 5. Priority: <prog-id-list>
                if mode_entry.startswith('VISITOR:'):
                    instrument = self.lookup_resource(mode_entry[8:].strip(), resource_type=ResourceType.INSTRUMENT)
                    instrument_run.setdefault(instrument, set()).add(row_date)

                elif mode_entry.startswith('PARTNER:'):
                    try:
                        partner = TimeAccountingCode[mode_entry[8:].strip()]
                    except KeyError as ex:
                        raise KeyError(f'{msg} has illegal time account {ex} in mode: {mode_entry}.')
                    # print(f'\t {row_date} {partner}')
                    FileBasedResourceService._add_dates_to_partdict(partner_blocks, partner, row_date)

                elif mode_entry.startswith('PV:'):
                    FileBasedResourceService._add_dates_to_dict(pv_programs, mode_entry[3:], row_date)

                elif mode_entry.startswith('CLASSICAL:'):
                    FileBasedResourceService._add_dates_to_dict(classical_programs, mode_entry[10:], row_date)

                elif mode_entry.startswith('PRIORITY:'):
                    FileBasedResourceService._add_dates_to_dict(score_boost, mode_entry[9:], row_date)

                elif mode_entry != 'QUEUE':
                    raise ValueError(f'{msg} has illegal mode: {mode_entry}.')

            # Get the LGS status of the row. The default, if no LGS is specified, is False.
            lgs = str_to_bool(row[3].value)
            if not row[3].value:
                logger.warning(f'{msg} has no LGS entry. Using default value of No.')
                self._lgs[site][row_date] = False
            else:
                self._lgs[site][row_date] = lgs

            # Get the ToO status for the night...
            too = str_to_bool(row[4].value)
            if not row[4].value:
                logger.warning(f'{msg} has no ToO entry. Using default value of Yes.')
                self._too[site][row_date] = True
            else:
                self._too[site][row_date] = too

            # Determine instrument and WFS resources available.
            resources: Set[Resource] = set()

            # The site is available on the day, so add it to the resources.
            resources.add(site.resource)

            # The next five entries contain instrument ports.
            # Filter out any ports that are not empty.
            # We then have to check to see what the status of the instrument is for the night.
            instrument_names = {row[i].value for i in range(5, 10) if row[i].value is not None and row[i].value}

            for filename in instrument_names:
                # TODO: Temporary: GCAL does not have its own column. Unsure of how to handle this.
                if filename != 'GCAL' and filename not in instrument_column_mapping:
                    raise KeyError(f'{msg} contains instrument name with no column: {filename}.')
                try:
                    # GCAL does not have its own column, treat as always available
                    if filename == 'GCAL':
                        instrument_status = FileBasedResourceService._SCIENCE
                    else:
                        instrument_status = none_to_str(row[instrument_column_mapping[filename]].value).strip().upper()
                except IndexError:
                    # This happens if the row ends prematurely.
                    instrument_status = ''
                if instrument_status == FileBasedResourceService._SCIENCE:
                    resources.add(self.lookup_resource(filename, resource_type=ResourceType.INSTRUMENT))
                    # Check for GRACES if GMOS-N is available (validation mode only)
                    if filename == 'GMOS-N' and 'GRACES' in instrument_column_mapping:
                        graces_status = none_to_str(row[instrument_column_mapping['GRACES']].value).strip().upper()
                        if graces_status == FileBasedResourceService._SCIENCE:
                            resources.add(self.lookup_resource('GRACES', resource_type=ResourceType.INSTRUMENT))
                elif not instrument_status:
                    logger.warning(f'{msg} contains no instrument status for {filename}. '
                                   'Using default of Not Available.')
                elif instrument_status not in [FileBasedResourceService._NOT_AVAILABLE,
                                               FileBasedResourceService._ENGINEERING,
                                               FileBasedResourceService._CALIBRATION]:
                    raise ValueError(f'{msg} for instrument {filename} '
                                     f'contains illegal status: {instrument_status}.')

            # The remaining columns are WFS. Check the status and if available, add to resources.
            for idx, filename in wfs_columns.items():
                try:
                    wfs_status = none_to_str(row[idx].value).strip().upper()
                except IndexError:
                    # This happens if the row ends prematurely.
                    wfs_status = ''
                if wfs_status == FileBasedResourceService._SCIENCE:
                    resources.add(self.lookup_resource(filename, resource_type=ResourceType.WFS))
                elif not wfs_status or wfs_status:
                    logger.warning(f'{msg} for WFS {filename} contains no status. Using default of Not Available.')
                elif (wfs_status not in
                        [FileBasedResourceService._NOT_AVAILABLE, FileBasedResourceService._ENGINEERING]):
                    raise ValueError(f'{msg} for WFS {filename} contains illegal status: {wfs_status}.')

            # Add the resource data to the dates. Union returns a new set.
            self._resources[site][row_date] = self._resources[site].setdefault(row_date, set()).union(resources)

        # Determine the earliest and latest date for the site.
        self._earliest_date_per_site[site] = min(dates.union(self._blocked[site]))
        self._latest_date_per_site[site] = max(dates.union(self._blocked[site]))

        # Block out the blocked dates.
        for d in self._blocked[site]:
            s = self._positive_filters[site].setdefault(d, set())
            s.add(NothingFilter())

        # PV rules:
        for pv_program_id, pv_dates in pv_programs.items():
            pv_starting_date = min(pv_dates)

            # 1. PV programs get priority on the nights they are listed.
            for d in pv_dates:
                s = self._positive_filters[site].setdefault(d, set())
                s.add(ProgramPriorityFilter(frozenset({pv_program_id})))

            # 2. PV programs are not allowed in nights before they are first listed.
            pv_prohibited_dates = {d for d in dates if d < pv_starting_date}
            for d in pv_prohibited_dates:
                s = self._negative_filters[site].setdefault(d, set())
                s.add(ProgramPermissionFilter(frozenset({pv_program_id})))

        # Classical rules:
        for classical_program_id, classical_dates in classical_programs.items():
            for d in dates:
                # Classical programs can only be performed in their designated blocks.
                if d in classical_dates:
                    s = self._positive_filters[site].setdefault(d, set())
                    s.add(ProgramPriorityFilter(frozenset({classical_program_id})))
                else:
                    s = self._negative_filters[site].setdefault(d, set())
                    s.add(ProgramPermissionFilter(frozenset({classical_program_id})))

        # Priority rules:
        for priority_program_id, priority_dates in score_boost.items():
            # Priority is given to programs in the priority block.
            for d in priority_dates:
                s = self._positive_filters[site].setdefault(d, set())
                s.add(ProgramPriorityFilter(frozenset({priority_program_id})))

        # Partner rules: programs can only be done on the nights specified
        for partner_code, partner_dates in partner_blocks.items():
            for d in dates:
                if d in partner_dates:
                    s = self._positive_filters[site].setdefault(d, set())
                    s.add(TimeAccountingCodeFilter(codes=frozenset({partner_code})))
                else:
                    s = self._negative_filters[site].setdefault(d, set())
                    s.add(TimeAccountingCodeFilter(codes=frozenset({partner_code})))

        # Visitor instrument rules:
        for resource, resource_dates in instrument_run.items():
            # Priority is given to scheduling blocks using the resource.
            for d in resource_dates:
                s = self._positive_filters[site].setdefault(d, set())
                s.add(ResourcePriorityFilter(frozenset({resource})))

        # ToO rules:
        for d in dates:
            # Block ToOs on nights where they are not allowed.
            if not self._too[site][d]:
                s = self._negative_filters[site].setdefault(d, set())
                s.add(TooFilter())

        # LGS rules:
        for d in dates:
            # Block LGS on nights where they are not allowed
            if not self._lgs[site][d]:
                s = self._negative_filters[site].setdefault(d, set())
                s.add(LgsFilter())

    @staticmethod
    def _mirror_parser(r: List[str], _: Site) -> Set[str]:
        return {'Mirror'} | {i.strip().replace('+', '') for i in r}

    def _load_time_loss(self,
                        site: Site,
                        name: str,
                        site_dict: Dict[Site, Dict[date, Set[T]]],
                        constructor: Type[T]) -> None:
        """
        Load the data from the specified file.
        A common syntax is used for:
        * engineering tasks
        * weather closures
        * faults
        """
        path = self._subdir / name

        try:
            with open(path, 'r') as input_file:
                pattern = r'(\d{4}-\d{2}-\d{2})\s+((?:\d{1,2}:\d{2})|twi)\s+((?:\d{1,2}:\d{2})|twi)(?:\s+\[(.*?)\])?'
                entries = site_dict[site]

                for line_num, line in enumerate(input_file):
                    line = line.strip()

                    # Skip blank lines or lines that start with a #
                    if not line or line[0] == '#':
                        continue

                    match = re.match(pattern, line)
                    if not match:
                        logger.warning(f'Illegal line {name}@{line_num + 1}: "{line}"')
                        continue

                    local_date_str, start_time_str, end_time_str, description = match.groups()
                    # print(f'{local_date_str} {start_time_str} {end_time_str}')
                    local_night_date = datetime.strptime(local_date_str, '%Y-%m-%d').date()

                    # Determine the start and end times.
                    start_time = None
                    end_time = None

                    if start_time_str != 'twi':
                        start_time = datetime.strptime(start_time_str, '%H:%M').time()
                    if end_time_str != 'twi':
                        end_time = datetime.strptime(end_time_str, '%H:%M').time()

                    if start_time is None or end_time is None:
                        # We need the twilights in this case.
                        # Local time when we change the UT designation for a night, just to get the closest midnight
                        new_ut_time = time(14,0)
                        astropy_time = Time(datetime.combine(local_night_date, new_ut_time).astimezone(site.timezone))

                        # Get the twilights and localize them.
                        eve_twi, morn_twi = night_events(astropy_time, site.location, site.timezone)[3:5]
                        eve_twi = eve_twi.to_datetime(site.timezone)
                        morn_twi = morn_twi.to_datetime(site.timezone)
                        # print(f'{local_date_str} {astropy_time} {eve_twi} {morn_twi} {site.timezone}')

                        if start_time is None:
                            start_time = (eve_twi - timedelta(seconds=10)).time()
                        if end_time is None:
                            end_time = morn_twi.replace(second=0, microsecond=0).time()

                    # Localize the datetimes. If the end time is before the start time, it happens on the next day.
                    night_entries_set = entries.setdefault(local_night_date, set())
                    start_datetime = datetime.combine(local_night_date, start_time).replace(tzinfo=site.timezone)
                    end_datetime = datetime.combine(local_night_date, end_time).replace(tzinfo=site.timezone)

                    # Add a day to the start_time and end_time if either time is less than noon to indicate that
                    # the event actually ends on the morning of the local_night_date.
                    # It should still be inserted for the local_night_date but to be sorted correctly,
                    # it needs to have a start time on the next day.
                    if start_datetime.time() < FileBasedResourceService._noon:
                        start_datetime += FileBasedResourceService._day
                    if end_datetime.time() < FileBasedResourceService._noon:
                        end_datetime += FileBasedResourceService._day
                    if end_datetime < start_datetime:
                        raise RuntimeError(f'Problem parsing date information for {path.name}: L{line_num}, {line} ')

                    entry = constructor(site=site,
                                        start_time=start_datetime,
                                        end_time=end_datetime,
                                        description=description)
                    night_entries_set.add(entry)

        except FileNotFoundError:
            logger.error(f'Time loss file not available: {path}')

    def _load_faults(self, site: Site, name: str) -> None:
        """
        Load the faults from the specified file.
        """
        path = self._subdir / name

        try:
            with open(path, 'r') as input_file:
                pattern = r'FR-(\d+)\s+(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\s+([\d.]+)\s+\[([^\]]+)\]'
                faults = self._faults[site]

                for line_num, line in enumerate(input_file):
                    line = line.strip()

                    # Skip blank lines or lines that start with a # indicating a comment.
                    if not line or line[0] == '#':
                        continue

                    match = re.match(pattern, line)
                    if not match:
                        logger.warning(f'Illegal line {name}@{line_num + 1}: "{line}"')
                        continue

                    fr_id, local_datetime_str, duration_str, description = match.groups()
                    local_datetime = datetime.strptime(local_datetime_str, '%Y-%m-%d %H:%M:%S')
                    local_datetime = local_datetime.replace(tzinfo=site.timezone)
                    duration = timedelta(hours=float(duration_str))

                    # Determine the night of the fault report from the local datetime.
                    # If it is before noon, it belongs to the previous night.
                    if local_datetime.time() < time(hour=12):
                        night_date = local_datetime.date() - timedelta(days=1)
                    else:
                        night_date = local_datetime.date()

                    # Add the fault to the night.
                    # TODO: Right now, not sure how to handle faults in terms of Resources.
                    # TODO: Just specify the entire site as a resource for now, indicating that the site cannot be used
                    # TODO: for the specified period.
                    # TODO: Fix duration as per email discussion.
                    faults.setdefault(night_date, set())
                    fault = Fault(site=site,
                                  start_time=local_datetime,
                                  end_time=local_datetime + duration,
                                  description=f'FR-{fr_id}: {description}')
                    faults[night_date].add(fault)
        except FileNotFoundError:
            logger.error(f'Faults file not available: {path}')

    def load_files(self,
                   site: Site,
                   fpu_to_barcodes_file: str,
                   fpus_data: Union[str, BytesIO],
                   gratings_data: Union[str, BytesIO],
                   faults_data: Union[str, BytesIO],
                   eng_tasks_data: Union[str, BytesIO],
                   weather_closure_data: Union[str, BytesIO],
                   spreadsheet_file: str) -> None:
        """
        Load all files necessaries to the correct functioning of the ResourceManager.
        """
        # Load the mappings from the ITCD FPU values to the barcodes.
        logger.debug(f'Reading FPU barcode data for {site}.')
        self._load_fpu_to_barcodes(site, fpu_to_barcodes_file)
        logger.debug(f'Done reading FPU barcode data for {site}.')

        # Load the FPUrs.
        # This will put both the IFU and the FPU barcodes available on a given date as Resources.
        # Note that for the IFU, we need to convert to a barcode, which is a Resource.
        # This is a bit problematic since we expect a list of strings of Resource IDs, so we have to take its ID.
        logger.debug(f'Reading IFU-FPU barcode data for {site}.')
        self._load_csv(site,
                       self._itcd_fpu_to_barcode_parser,
                       fpus_data,
                       resource_type=ResourceType.FPU)
        logger.debug(f'Done reading IFU-FPU barcode data for {site}.')

        # Load the gratings.
        # This will put the mirror and the grating names available on a given date as Resources.
        # TODO: Check Mirror vs. MIRROR. Seems like GMOS uses Mirror.
        logger.debug(f'Reading gratings data for {site}.')
        self._load_csv(site,
                       self._mirror_parser,
                       gratings_data,
                       resource_type=ResourceType.DISPERSER)
        logger.debug(f'Done reading gratings data for {site}.')

        # Process the spreadsheet information for instrument, mode, and LGS settings.
        logger.debug(f'Reading instrument data for {site}.')
        self._load_instrument_data(site, spreadsheet_file)
        logger.debug(f'Done reading instrument data for {site}.')

        logger.debug(f'Reading fault data for {site}.')
        self._load_time_loss(site, faults_data, self._faults, Fault)
        logger.debug(f'Done reading fault data for {site}.')

        logger.debug(f'Reading engineering task data for {site}.')
        self._load_time_loss(site, eng_tasks_data, self._eng_tasks, EngineeringTask)
        logger.debug(f'Done reading engineering task data for {site}.')

        logger.debug(f'Reading weather closure data for {site}.')
        self._load_time_loss(site, weather_closure_data, self._weather_closures, WeatherClosure)
        logger.debug(f'Done reading weather closure data for {site}.')
