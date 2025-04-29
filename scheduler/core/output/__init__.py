# Copyright (c) 2016-2024 Association of Universities for Research in Astronomy, Inc. (AURA)
# For license information see LICENSE or https://opensource.org/licenses/BSD-3-Clause

from datetime import timedelta
import gzip
import json
import os
import pickle
import sys
from typing import List, Dict

from astropy import units as u
from lucupy.minimodel import Atom, Group, Observation, ObservationClass, Program, Site
from openpyxl import Workbook
from pandas import DataFrame
import numpy as np

from scheduler.core.components.collector import Collector
from scheduler.core.components.nighteventsmanager import NightEventsManager
from scheduler.core.plans import Plans
from scheduler.core.programprovider.abstract import ProgramProvider
from scheduler.core.programprovider.ocs import OcsProgramProvider
from scheduler.core.calculations.selection import Selection


__all__ = [
    'print_program_from_provider',
    'print_collector_info',
    'print_atoms_for_observation',
    'atoms_to_sheet',
    'print_plans',
    'plans_table',
    'pickle_plans',
    'pickle_selection',
]


def print_program_from_provider(filename=os.path.join('data', 'GN-2018B-Q-101.json.gz'),
                                provider: ProgramProvider = OcsProgramProvider) -> None:
    """
    Using a specified JSON file and a ProgramProvider, read in the program
    and print it.
    """
    with gzip.open(filename, 'r') as f:
        data = json.loads(f.read())

    program = provider.parse_program(data['PROGRAM_BASIC'])
    program.show()


def print_collector_info(collector: Collector) -> None:
    sys.stderr.flush()
    # Output some information.
    print('Pre-Collector / Collector running from:')
    print(f'   start time:       {collector.start_vis_time}')
    print(f'   end time:         {collector.end_vis_time}')
    print(f'   time slot length: {collector.time_slot_length.to(u.min)}')

    # Print out sampled calculation for every hour as there are far too many results to print in full.
    time_grid = collector.time_grid
    for site in collector.sites:
        nc = collector.night_configurations(site, np.arange(collector.num_nights_calculated))
        print(f'\n\n+++++ NIGHT EVENTS FOR {site.name} +++++')
        night_events = NightEventsManager.get_night_events(collector.time_grid, collector.time_slot_length, site)
        for idx, jday in enumerate(time_grid):
            start = night_events.local_times[idx][0]
            end = night_events.local_times[idx][-1]
            num_slots = len(night_events.times[idx])
            print(f'* DAY {idx}: {start} to {end}, {num_slots} time slots.')
            print(f'\tmidnight:              {night_events.midnight[idx]}')
            print(f'\tsunset:                {night_events.sunset[idx]}')
            print(f'\tsunrise:               {night_events.sunrise[idx]}')
            print(f'\t12° eve twilight:      {night_events.twilight_evening_12[idx]}')
            print(f'\t12° mor twilight:      {night_events.twilight_morning_12[idx]}')
            print(f'\tmoonrise:              {night_events.moonrise[idx]}')
            print(f'\tmoonset:               {night_events.moonset[idx]}')
            print(f"\tInstruments available: {', '.join(sorted(r.id for r in nc[idx].resources))}")
    sys.stdout.flush()


def print_atoms_for_observation(observation: Observation) -> None:
    for atom in observation.sequence:
        print(f'\t{atom}')


def atoms_to_sheet(dt: Program | Observation | Group) -> None:
    """
    Print out the atoms in a program or observation to a spreadsheet.
    """
    wb = Workbook()
    ws = wb.active
    ws.append(['id', 'exec_time', 'prog_time', 'part_time', 'observed', 'qa_state', 'guide_state'])

    def save_to_sheet(curr_atom: Atom):
        ws.cell(column=1, row=curr_atom.id + 1, value=curr_atom.id)
        ws.cell(column=2, row=curr_atom.id + 1, value=curr_atom.exec_time.total_seconds())
        ws.cell(column=3, row=curr_atom.id + 1, value=curr_atom.prog_time.total_seconds())
        ws.cell(column=4, row=curr_atom.id + 1, value=curr_atom.part_time.total_seconds())
        ws.cell(column=5, row=curr_atom.id + 1, value=curr_atom.observed)
        ws.cell(column=6, row=curr_atom.id + 1, value=curr_atom.qa_state.value)
        ws.cell(column=7, row=curr_atom.id + 1, value=curr_atom.guide_state)

    # TODO: Output for larger formats(e.g Program) not required but might be good to have.
    if isinstance(dt, Program):
        for obs in dt.observations():
            if obs.obs_class in [ObservationClass.SCIENCE, ObservationClass.PARTNERCAL]:
                for atom in obs.sequence:
                    ws.title = f'{obs.id}'
                    save_to_sheet(atom)
                ws = wb.create_sheet()
                ws.append(['id', 'exec_time', 'prog_time', 'part_time', 'observed', 'qa_state', 'guide_state'])
        wb.save(f'{dt.id}.xlsx')

    elif isinstance(dt, Observation):
        print(f'Observation: {dt.id}')
        for atom in dt.sequence:
            save_to_sheet(atom)
        wb.save(f'{dt.id}.xlsx')
    elif isinstance(dt, Group):
        raise NotImplementedError
    else:
        raise ValueError(f'Unsupported type: {type(dt)}')


def print_plans(all_plans: List[Plans]) -> None:
    """
    Print out the visit plans.
    """
    sys.stderr.flush()
    print(f'\n\n+++++ NIGHT {all_plans.night_idx + 1} +++++')
    for plan in sorted(all_plans, key=lambda x: x.site.name):
        if max_score_length := max((len(f'{visit.score:8.2f}') for visit in plan.visits), default=0):
            print(f'Plan for site: {plan.site.name}')
            print(f'Conditions: IQ {plan.conditions.iq} CC {plan.conditions.cc}')
            print(f'{"Execution time":{36}}   {"ObsID":{20}} {"Score":>{max_score_length}}'
                  '  StartAtom    EndAtom  StartSlot    EndSlot   NumSlots')
            for visit in plan.visits:
                start_time_str = visit.start_time.strftime('%Y-%m-%d %H:%M')
                end_time_str = (visit.start_time + timedelta(minutes=visit.time_slots)).strftime('%Y-%m-%d %H:%M')
                print(f'{start_time_str} to {end_time_str}   {visit.obs_id.id:20} {visit.score:8.2f}  '
                      f'{visit.atom_start_idx:9d}  {visit.atom_end_idx:9d}  {visit.start_time_slot:9d}  '
                      f' {(visit.start_time_slot + visit.time_slots - 1):9d} {visit.time_slots:9d}')
        else:
            print(f'Empty plan for site: {plan.site.name}')


def plans_table(all_plans: List[Plans]) -> List[Dict[Site, DataFrame]]:
    sys.stderr.flush()
    per_night = []
    for plans in all_plans:
        per_site = {}
        for plan in plans:
            new_entry = {'Start': [v.start_time for v in plan.visits],
                         'End': [v.start_time+timedelta(minutes=v.time_slots) for v in plan.visits],
                         'Observation': [v.obs_id.id for v in plan.visits],
                         'Class': [v.obs_class.name for v in plan.visits],
                         'Atom start': [v.atom_start_idx for v in plan.visits],
                         'Atom end': [v.atom_end_idx for v in plan.visits],
                         'Length': [v.time_slots for v in plan.visits],
                         'Score': [v.score for v in plan.visits],
                         'Instrument': [v.instrument.id for v in plan.visits]}
            df = DataFrame(new_entry)
            per_site[plan.site] = df
        per_night.append(per_site)

    return per_night


def pickle_plans(plans_to_pickle: List[Dict[Site, DataFrame]],
                 path: str,
                 start: str = '',
                 end: str = '') -> None:
    with open(f'{path}/plans_{start}_to_{end}.pickle', 'wb') as f:
        pickle.dump(plans_to_pickle, f)


def pickle_selection(selection_to_pickle: Selection,
                     path: str,
                     night: str = '') -> None:
    with open(f'{path}/selection_night{night}.pickle', 'wb') as f:
        pickle.dump(selection_to_pickle, f)


def print_observations(group, print_targ=False, print_atoms=False):
    """Print details of the observations in a group"""
    for obs in group.observations():
        print(f'\t\t\t Obs: {obs.id.id} {obs.priority.name} IQ:{obs.constraints.conditions.iq} '
              f'CC:{obs.constraints.conditions.cc} {obs.exec_time()} {obs.obs_class.name:12} {obs.total_used()} '\
              f'{obs.status.name} Pre-imaging:{obs.preimaging}') # {min(program.target_info[obs.id][0].airmass):5.2f}')
        if print_targ:
            for target in obs.targets:
                try:
                    ra = target.ra
                    dec = target.dec
                except Exception:
                    ra = None
                    dec = None
                print(f'\t\t\t\t {target.name} {target.type} {ra} {dec}')
        if print_atoms:
            for atom in obs.sequence:
                print(
                    f'\t\t\t\tAtom: {atom.id} {atom.exec_time} {atom.prog_time} {atom.part_time} {atom.observed} '
                    f'{atom.qa_state.name} {atom.program_used} {atom.partner_used}')


def print_selection(selection, print_obs=False, print_targ=False, print_atoms=False):
    """Print details of a selection, currently appropriate for Validation mode (not recursive)"""
    # print(f'Night indic   es: {selection.night_indices}')
    night_idx = selection.night_indices[0]
    for p in selection.program_info.values():
        print(f'Program: {p.program.id.id}  mean user priority:{p.program.mean_priority():5.2f}')
    #     for targ in p.target_info:
    #         print(f'{targ.id}')
        for g in p.group_data_map.values():
            print(f'\t Group: {g.group.unique_id.id} {g.group.priority().name} ObsGroup:{g.group.is_observation_group()} '\
                  f'SchedGroup:{g.group.is_scheduling_group()} AndGroup:{g.group.is_and_group()} '\
                  f'Max score: {np.max(g.group_info.scores[night_idx]):7.2f} Exec time: {g.group.exec_time()}')
            if g.group.is_scheduling_group():
                for subgroup in g.group.children:
                    # print(f'\t\t {subgroup.unique_id}')
                    if subgroup.unique_id in p.group_data_map:
                        sg = p.group_data_map[subgroup.unique_id]
                        max_score = np.max(sg.group_info.scores[night_idx])
                    else:
                        max_score = 0.0
                    print(f'\t\t {subgroup.unique_id.id} {subgroup.priority().name} {subgroup.exec_time()} {subgroup.is_observation_group()}\
                    {max_score:7.2f}')
                    if print_obs:
                        print_observations(subgroup, print_targ=print_targ, print_atoms=print_atoms)
            elif print_obs:
                print_observations(g.group, print_targ=print_targ, print_atoms=print_atoms)


def print_schedulable_groups(selection, print_obs=False, print_targ=False, print_atoms=False):
    """Print details of the schedulable groups, currently appropriate for Validation mode (not recursive)"""

    night_idx = selection.night_indices[0]
    for group_data in selection.schedulable_groups.values():
        p = selection.program_info[group_data.group.program_id]
        constraints = group_data.group.constraints()
        # print(constraints)
        if group_data.group.unique_id in p.group_data_map:
            sg = p.group_data_map[group_data.group.unique_id]
            max_score = np.max(sg.group_info.scores[night_idx])
        else:
            max_score = 0.0
        print(f'{group_data.group.unique_id.id} IQ:{constraints[0].conditions.iq} CC:{constraints[0].conditions.cc} '
              f'Max score: {max_score:7.2f} Exec time: {group_data.group.exec_time()}')
        print(f'\t{group_data.group.required_resources()}')
        if group_data.group.is_scheduling_group():
            for subgroup in group_data.group.children:
                constraints = subgroup.constraints()
                # print(len(constraints))
                if subgroup.unique_id in p.group_data_map:
                    sg = p.group_data_map[subgroup.unique_id]
                    max_score = np.max(sg.group_info.scores[night_idx])
                else:
                    max_score = 0.0
                print(f'\t {subgroup.id.id} IQ:{constraints[0].conditions.iq} CC:{constraints[0].conditions.cc} '
                      f'Max score:{max_score:7.2f}  Exec time: {subgroup.exec_time()}')
                print(f'\t\t{subgroup.required_resources()}')
                if print_obs:
                    print_observations(subgroup, print_targ=False, print_atoms=print_atoms)
        elif print_obs:
            print_observations(group_data.group, print_targ=print_targ, print_atoms=print_atoms)
