# Copyright (c) 2016-2022 Association of Universities for Research in Astronomy, Inc. (AURA)
# For license information see LICENSE or https://opensource.org/licenses/BSD-3-Clause

import gzip
import json
import os
import sys
from typing import Dict, FrozenSet, Optional, NoReturn, Sequence, TypeVar

import matplotlib.pyplot as plt
import numpy as np
import requests
from lucupy.minimodel import ObservationClass, QAState
from openpyxl import Workbook
from openpyxl import load_workbook
from scipy.signal import find_peaks

from definitions import ROOT_DIR

T = TypeVar('T')

fpuinst = {'GSAOI': 'instrument:utilityWheel', 'GPI': 'instrument:observingMode', 'Flamingos2': 'instrument:fpu',
           'NIFS': 'instrument:mask', 'GNIRS': 'instrument:slitWidth', 'GMOS-N': 'instrument:fpu',
           'GMOS-S': 'instrument:fpu', 'NIRI': 'instrument:mask'}

gpi_filter_wav = {'Y': 1.05, 'J': 1.25, 'H': 1.65, 'K1': 2.05, 'K2': 2.25}
nifs_filter_wav = {'ZJ': 1.05, 'JH': 1.25, 'HK': 2.20}


def find_filter(input_filter: str, filter_dict: Dict[str, float]) -> Optional[str]:
    """
    Match input string with filter list (in dictionary).
    """
    return next((x for x in list(filter_dict.keys()) if x in input_filter), None)


def search_list(val: Sequence[T], alist: Sequence[T]) -> T:
    """
    Search for existence of val in any element of alist.
    """
    return any(elem for elem in alist if val in elem)


def short_id(id: str) -> str:
    """
    Return short form of obs_id or data label.
    """
    id_split = id.split('-')
    id_out = id_split[0][1] + id_split[1][2:5] + '-' + id_split[2] + '-' + id_split[3] + '[' + id_split[4] + ']'
    if len(id_split) == 6:
        id_out += '-' + id_split[5]
    return id_out


def odb_json(prog_id: str,
             path: Optional[str] = None,
             overwrite: bool = False):
    """
    Download json of ODB program information

    Parameters
        progid:     Program ID of program to extract
        path:       Path for json files
        overwrite:  Overwrite any existing json files

    Return
        json_result: JSON query result as a list of dictionaries
    """
    if not prog_id:
        raise ValueError('Program id not given.')

    file = prog_id + '.json.gz'
    if path is not None and not overwrite and os.path.exists(os.path.join(path, file)):
        with gzip.open(os.path.join(path, file), 'r') as fin:
            json_bytes = fin.read()

        json_str = json_bytes.decode('utf-8')
        json_result = json.loads(json_str)
    else:
        response = requests.get(
            'http://gnodbscheduler.hi.gemini.edu:8442/programexport?id=' + prog_id)
        try:
            response.raise_for_status()
        except requests.exceptions.HTTPError as exc:
            raise exc
        else:
            json_result = response.json()
            if path is not None and (overwrite or not os.path.exists(os.path.join(path, file))):
                json_str = json.dumps(json_result, indent=2)
                with gzip.open(os.path.join(path, file), 'wb') as fout:
                    fout.write(json_str.encode('utf-8'))

    return json_result


def obsmode(config: Dict[str, str]) -> str:
    """
    Determine the observation mode (e.g. imaging, longslit, mos, ifu, etc.
    """
    mode = 'unknown'
    if search_list('GMOS', config['inst']):
        if 'MIRROR' in config['disperser']:
            mode = 'imaging'
        elif search_list('arcsec', config['fpu']):
            mode = 'longslit'
        elif search_list('IFU', config['fpu']):
            mode = 'ifu'
        elif 'CUSTOM_MASK' in config['fpu']:
            mode = 'mos'
    elif config['inst'] in ["GSAOI", "'Alopeke", "Zorro"]:
        mode = 'imaging'
    elif config['inst'] in ['IGRINS', 'MAROON-X']:
        mode = 'longslit'
    elif config['inst'] in ['GHOST', 'MAROON-X', 'GRACES', 'Phoenix']:
        mode = 'xd'
    elif config['inst'] == 'Flamingos2':
        if search_list('LONGSLIT', config['fpu']):
            mode = 'longslit'
        if search_list('FPU_NONE', config['fpu']) \
                and search_list('IMAGING', config['disperser']):
            mode = 'imaging'
    elif config['inst'] == 'NIRI':
        if search_list('NONE', config['disperser']) and search_list('MASK_IMAGING', config['fpu']):
            mode = 'imaging'
    elif config['inst'] == 'NIFS':
        mode = 'ifu'
    elif config['inst'] == 'GNIRS':
        if search_list('mirror', config['disperser']):
            mode = 'imaging'
        elif search_list('XD', config['disperser']):
            mode = 'xd'
        else:
            mode = 'longslit'
    elif config['inst'] == 'GPI':
        if search_list('CORON', config['fpu']):
            mode = 'coron'
        elif search_list('NRM', config['fpu']):
            mode = 'nrm'
        elif search_list('DIRECT', config['fpu']):
            mode = 'imaging'
        else:
            mode = 'ifu'

    return mode


def guide_state(step):
    """
    Determine if guiding is on/off for a sequence step.
    """
    return any(key for key in list(step.keys()) if 'guideWith' in key and step[key] == 'guide')


def select_qastate(states: Sequence[QAState]) -> str:
    """
    Return the qastate based on precedence.
    This requires special handling, so we leave it instead of just using enum comparison.

    states: list of observe states from the ODB extractor obsLog
    """
    # Precedence order for observation classes.
    qastate_order = ['NONE', 'UNDEFINED', 'FAIL', 'USABLE', 'PASS']

    # Set the qastate for the entire observation based on precedence
    for state in qastate_order:
        if state in states:
            return state
    return ''


def select_obsclass(classes: Sequence[ObservationClass]) -> str:
    """
    Return the obsclass based on precedence
    This requires special handling, so we leave it instead of just using enum comparison.

    classes: list of observe classes from the ODB extractor
    """
    # Precedence order for observation classes.
    obsclass_order = ['SCIENCE', 'PROGCAL', 'PARTNERCAL', 'ACQ', 'ACQCAL', 'DAYCAL']

    # Set the obsclass for the entire observation based on obsclass precedence
    for oclass in obsclass_order:
        if oclass in classes:
            return oclass
    return ''


def autocorr_lag(x, plot=False):
    """
    Test for patterns with auto-correlation.
    """
    # Auto correlation
    result = np.correlate(x, x, mode='full')
    corrmax = np.max(result)
    if corrmax != 0.0:
        result = result / corrmax
    if plot:
        plt.plot(result[result.size // 2:])

    # Pattern offset using first prominent peak
    peaks, prop = find_peaks(result[result.size // 2:], height=(0, None), prominence=(0.25, None))
    return peaks[0] if any(peaks) else 0


def find_atoms(observation, verbose=False, ws=None, fid=sys.stdout):
    """Analyze a json observing sequence from the ODB and define atoms."""
    observe_types = frozenset(['FLAT', 'ARC', 'DARK', 'BIAS'])

    classes = []
    guiding = []
    qastates = []
    atoms = []
    natom = 0

    # Make dictionary out of obslog to get QA state
    obslog = {}
    datalabels = []
    if 'obsLog' in observation.keys():
        for log_entry in observation['obsLog']:
            obslog[log_entry['label']] = {'qaState': log_entry['qaState'], 'filename': log_entry['filename']}
        datalabels = list(obslog.keys())

    # Sequence analysis
    sequence = observation['sequence']
    nsteps = len(sequence)

    # First pass to analyze offsets and exptimes/coadds
    do_not_split = False
    exptimes = []
    coadds = []
    qoffsets = []
    poffsets = []
    qoffsets_sky = []
    poffsets_sky = []
    config = {'inst': '', 'fpu': [], 'disperser': [], 'filter': [], 'wavelength': []}

    for idx, step in enumerate(sequence):
        step_keys = list(step.keys())

        inst = step['instrument:instrument']
        # print(inst, fpuinst[inst])
        if inst == 'Visitor Instrument':
            inst = step['instrument:name'].split(' ')[0]
            fpu = 'None' if inst in ["'Alopeke", "Zorro"] else inst
        else:
            fpu = step[fpuinst[inst]]
        config['inst'] = inst
        config['fpu'].append(fpu)

        if 'instrument:disperser' in step_keys:
            disperser = step['instrument:disperser']
        elif inst in ['IGRINS', 'MAROON-X']:
            disperser = inst
        else:
            disperser = 'None'
        if inst == 'GNIRS':
            if step['instrument:acquisitionMirror'] == 'in' and \
                    step['instrument:decker'] == 'acquisition':
                disperser = 'mirror'
            else:
                disperser = disperser.strip('grating') + step['instrument:crossDispersed']
        elif inst == 'Flamingos2' and fpu == 'FPU_NONE':
            if step['instrument:decker'] == 'IMAGING':
                disperser = step['instrument:decker']
        config['disperser'].append(disperser)

        if 'instrument:filter' in step_keys:
            filt = step['instrument:filter']
        elif inst == 'GPI':
            filt = find_filter(fpu, gpi_filter_wav)
        else:
            if inst == 'GNIRS':
                filt = 'None'
            else:
                filt = 'Unknown'
        if inst == 'NIFS' and 'Same as Disperser' in filt:
            for filt in list(nifs_filter_wav.keys()):
                if disperser[0] in filt:
                    filt = filt
                    break
        config['filter'].append(filt)

        if inst == 'GPI':
            wavelength = gpi_filter_wav[filt]
        else:
            wavelength = float(step['instrument:observingWavelength'])
        config['wavelength'].append(wavelength)

        # Just want exposures on sky for dither pattern analysis
        p = 0.0
        q = 0.0

        if step['observe:observeType'].upper() not in observe_types:
            p = float(step['telescope:p']) if 'telescope:p' in step_keys else 0.0
            poffsets_sky.append(p)

            q = float(step['telescope:q']) if 'telescope:q' in step_keys else 0.0
            qoffsets_sky.append(q)

        ncoadds = int(step['observe:coadds']) if 'observe:coadds' in step_keys else 1
        poffsets.append(p)
        qoffsets.append(q)

        exptimes.append(float(step['observe:exposureTime']))
        coadds.append(ncoadds)

    mode = obsmode(config)
    if verbose:
        print(config)
        print(mode)

    if config['inst'] == 'GPI':
        do_not_split = True

    # Analyze sky offset patterns using auto-correlation
    # The lag is the length of any pattern, 0 means no repeating pattern
    plag = 0
    qlag = 0
    if do_not_split:
        offset_lag = nsteps
    else:
        if len(poffsets_sky) > 1:
            plag = autocorr_lag(np.array(poffsets_sky))
        if len(qoffsets_sky) > 1:
            qlag = autocorr_lag(np.array(qoffsets_sky))

        # Special cases
        if plag == 0 and qlag == 0 and len(qoffsets_sky) == 4:
            # single ABBA pattern, which the auto-correlation won't find
            if qoffsets_sky[0] == qoffsets_sky[3] and qoffsets_sky[1] == qoffsets_sky[2]:
                qlag = 4
        elif len(qoffsets_sky) == 2:
            # If only two steps, put them together, might be AB, also silly to split only two steps
            qlag = 2

        offset_lag = qlag
        if plag > 0 and plag != qlag:
            offset_lag = 0

        # Turn off offset groups if the length is longer than some limit
        # if offset_lag > 0 and plag > 0 and qlag > 0 and \
        #     np.sum(np.asarray(exptimes[0:offset_lag]) * np.asarray(coadds[0:offset_lag])) > 600.:
        #     offset_lag = 0

    print('Offset lags: ', plag, qlag, offset_lag, file=fid)

    # Group by changes in exptimes/coadds?
    exptime_groups = False
    # if len(uniquelist(exptimes)) > 1 or len(uniquelist(coadds)) > 1:
    #     exptime_groups = True

    # Group by changes in exptimes/coadds?
    # Write results to the Excel worksheet if given
    columns = ['datalab', 'class', 'type', 'inst', 'exec_time', 'exptime', 'coadds', 'fpu', 'filter',
               'disperser', 'wavelength', 'p', 'q', 'guiding', 'qa_state', 'atom']
    row = 1

    if ws is not None:
        for idx, col in enumerate(columns):
            ws.cell(column=idx + 1, row=row, value=col)
        row += 1

    npattern = offset_lag
    noffsets = 0
    idx_prevobj = -1
    idx = 0
    for idx, step in enumerate(sequence):
        nextatom = False

        datalab = step['observe:dataLabel']
        if datalab in datalabels:
            qastate = obslog[datalab]['qaState']
        else:
            qastate = 'NONE'

        observe_class = step['observe:class']

        step_time = step['totalTime'] / 1000.

        atomstr = 'Atom for: '
        # Any wavelength/filter change is a new atom
        if idx == 0 or (idx > 0 and config['wavelength'][idx] != config['wavelength'][idx - 1]):
            nextatom = True
            atomstr += 'wavelength, '

        # A change in exposure time or coadds is a new atom for science exposures
        if step['observe:observeType'].upper() not in ['FLAT', 'ARC', 'DARK', 'BIAS']:
            if (observe_class.upper() == 'SCIENCE' and idx > 0 and
                    (exptimes[idx] != exptimes[idx_prevobj] or coadds[idx] != coadds[idx_prevobj])):
                nextatom = True
                atomstr += 'exposure time change, '

            # Offsets - a new offset pattern is a new atom
            # print('npattern: ', npattern)
            if offset_lag != 0 or not exptime_groups:
                # For NIR imaging, need to have at least two offset positions if no repeating pattern
                # New atom after every 2nd offset (noffsets is odd)
                if mode == 'imaging' and offset_lag == 0 and all(w > 1.0 for w in config['wavelength']):
                    if idx == 0:
                        noffsets += 1
                    else:
                        if poffsets[idx] != poffsets[idx_prevobj] or qoffsets[idx] != qoffsets[idx_prevobj]:
                            noffsets += 1
                    if noffsets % 2 == 1:
                        nextatom = True
                        atomstr += 'offset pattern'
                else:
                    npattern -= 1
                    if npattern < 0:
                        nextatom = True
                        atomstr += 'offset pattern'
                        npattern = offset_lag - 1

            idx_prevobj = idx

        # New atom?
        if nextatom:
            # Get class, qastate, guiding for previous atom
            if natom > 0:
                atoms[-1]['qa_state'] = select_qastate(qastates)
                if atoms[-1]['qa_state'] != 'NONE':
                    atoms[-1]['observed'] = True
                atoms[-1]['class'] = select_obsclass(classes)
                if verbose:
                    print('QA states: ', qastates, file=fid)
                    print('Classes: ', classes, file=fid)
                    print('Guiding: ', guiding, file=fid)
                atoms[-1]['guide_state'] = any(guiding)
                atoms[-1]['wavelength'] = config['wavelength'][idx]
                atoms[-1]['required_resources']['inst'] = config['inst']
                atoms[-1]['required_resources']['filter'] = config['filter'][idx]
                atoms[-1]['required_resources']['disperser'] = config['disperser'][idx]
                atoms[-1]['required_resources']['fpu'] = config['fpu'][idx]

                # Print basic atom info
                print(" \t exec_time: {:7.2f}, prog_time: {:7.2f}, part_time: {:7.2f}, guide_state: {}".
                      format(atoms[-1]['exec_time'], atoms[-1]['prog_time'], atoms[-1]['part_time'],
                             atoms[-1]['guide_state']), file=fid)

            # New atom entry
            print(atomstr, file=fid)
            natom += 1
            atoms.append({'id': natom, 'exec_time': 0.0, 'prog_time': 0.0, 'part_time': 0.0,
                          'class': 'NONE', 'observed': False, 'qa_state': 'NONE', 'guide_state': False,
                          'wavelength': 0.0,
                          'required_resources': {'inst': 'NONE', 'filter': 'NONE', 'disperser': 'NONE',
                                                 'fpu': 'NONE'}})
            classes = []
            guiding = []
            qastates = []
            if step['observe:observeType'].upper() in ['FLAT', 'ARC', 'DARK', 'BIAS'] \
                    and npattern == 0:
                npattern = offset_lag
            noffsets = 1

        # Update atom
        qastates.append(QAState[qastate.upper()])
        classes.append(ObservationClass[observe_class.upper()])
        guiding.append(guide_state(step))

        atoms[-1]['exec_time'] += step_time

        atomlabel = natom
        if 'partnerCal' in observe_class:
            atomlabel *= 10
            atoms[-1]['part_time'] += step_time
        else:
            atoms[-1]['prog_time'] += step_time

        print('{:17} {:3} {:7.2f} {:3d} {:10} {:12} {:10} {:12} {:6.4f} {:8.2f} {:8.2f} {:1} {:3d}'.format(
            short_id(datalab),
            observe_class[0:3], exptimes[idx], coadds[idx], config['inst'], config['fpu'][idx],
            config['filter'][idx], config['disperser'][idx], config['wavelength'][idx], poffsets[idx],
            qoffsets[idx], guiding[-1], atomlabel),
            file=fid)
        if ws is not None:
            data = [datalab, observe_class.upper(), step['observe:observeType'].upper(), config['inst'],
                    step_time, exptimes[idx], coadds[idx], config['fpu'][idx],
                    config['filter'][idx], config['disperser'][idx], config['wavelength'][idx], poffsets[idx],
                    qoffsets[idx], guiding[-1], qastate.upper(), atomlabel]
            for col_idx, col_data in enumerate(data):
                ws.cell(column=col_idx + 1, row=row, value=col_data)
            row += 1

    # Get class/state for last atom
    if natom > 0:
        atoms[-1]['qa_state'] = select_qastate(qastates)
        if atoms[-1]['qa_state'] != 'NONE':
            atoms[-1]['observed'] = True
        atoms[-1]['class'] = select_obsclass(classes)
        if verbose:
            print('QA states: ', qastates, file=fid)
            print('Classes: ', classes, file=fid)
            print('Guiding: ', guiding, file=fid)
        atoms[-1]['guide_state'] = any(guiding)
        atoms[-1]['wavelength'] = config['wavelength'][idx]
        atoms[-1]['required_resources']['inst'] = config['inst']
        atoms[-1]['required_resources']['filter'] = config['filter'][idx]
        atoms[-1]['required_resources']['disperser'] = config['disperser'][idx]
        atoms[-1]['required_resources']['fpu'] = config['fpu'][idx]

        # Print basic atom info
        print(" \t exec_time: {:7.2f}, prog_time: {:7.2f}, part_time: {:7.2f}, guide_state: {}".
              format(atoms[-1]['exec_time'], atoms[-1]['prog_time'], atoms[-1]['part_time'],
                     atoms[-1]['guide_state']), file=fid)

    return atoms


def group_proc(group,
               sel_obs_class: FrozenSet = frozenset(['SCIENCE', 'PROGCAL', 'PARTNERCAL', 'ACQ', 'ACQCAL', 'DAYCAL']),
               sel_obs_status: FrozenSet = frozenset(['PHASE_2', 'FOR_REVIEW', 'IN_REVIEW', 'FOR_ACTIVATION', 'ON_HOLD',
                                                      'READY', 'ONGOING', 'OBSERVED', 'INACTIVE']),
               fid=sys.stdout, wb=None, verbose=False) -> NoReturn:
    """
    Process observations within groups.
    """
    ws = None
    obs_num = []
    for item in list(group.keys()):
        # obs_id = ''
        if 'OBSERVATION' in item:
            # obs_id = group[item]['observationId']
            obs_num.append(int(item.split('-')[1]))

    if len(obs_num) > 0:
        isrt = np.argsort(obs_num)
        for ii in isrt:
            # obs_program_used = 0.0
            # obs_partner_used = 0.0
            item = 'OBSERVATION_BASIC-' + str(obs_num[ii])
            # obs_id = program[prog][group][item]['sequence'][0]['ocs:observationId']
            obs_id = group[item]['observationId']
            print(f'{obs_num[ii], obs_id}', file=fid)
            obs_class = group[item]['obsClass'].upper()
            # phase2stat = group[item]['phase2Status'].upper()
            obs_stat = group[item]['obsStatus'].upper()
            # print(obs_class, phase2stat, obs_stat)
            if obs_class in sel_obs_class and obs_stat in sel_obs_status:
                if wb is not None:
                    ws = wb.create_sheet(title=obs_id.split('-')[-1])
                # ws['A1'] = obsid

                # Atoms in each sequence
                # TODO: Return value is not used but if this is not called, obs_id is not set and an error occurs.
                atoms = find_atoms(group[item], verbose=verbose, ws=ws, fid=fid)
                # Summary of atoms
                #                 classes = []
                #                 qastates = []
                #                 for atom in atoms:
                #                     print('Atom ', atom['id'])
                #                     for key in atom.keys():
                #                         print(f" \t {key}: {atom[key]}")
                #                         if key == 'class':
                #                             classes.append(atom[key])
                #                         if key == 'qa_state':
                #                             qastates.append(atom[key])
                #                             if atom[key].upper() == 'PASS':
                #                                 obs_program_used += atom['prog_time']
                #                                 obs_partner_used += atom['part_time']
                #                 obsclass = select_obsclass(classes)
                #                 print(f" Obsclass: {obsclass}")
                #                 obs_qastate = select_qastate(qastates)
                #                 print(f" QAstate (atoms): {obs_qastate}")
                #                 print(f" qaState (ODB): {group[item]['qaState']}")
                #                 if group[item]['qaState'].upper() == 'PASS':
                #                     if group[item]['obsClass'] in ['science', 'progCal']:
                #                         obs_program_used += float(group[item]['setupTime'])/1000.
                #                     elif group[item]['obsClass'] in ['partnerCal']:
                #                         obs_partner_used += float(group[item]['setupTime'])/1000.
                #                 print(f" program_used: {obs_program_used}")
                #                 print(f" partner_used: {obs_partner_used}")
                print('', file=fid)
            print('', file=fid)


def prog_proc(program,
              sel_obs_class: FrozenSet[str] = frozenset(['SCIENCE', 'PROGCAL', 'PARTNERCAL',
                                                         'ACQ', 'ACQCAL', 'DAYCAL']),
              sel_obs_status: FrozenSet[str] = frozenset(['PHASE_2', 'FOR_REVIEW', 'IN_REVIEW', 'FOR_ACTIVATION',
                                                          'ON_HOLD', 'READY', 'ONGOING', 'OBSERVED', 'INACTIVE']),
              fid=sys.stdout, xls=None) -> NoReturn:
    """
    Process top-level of program.
    """
    wb = None
    grpnum = []
    grplist = []
    for prog in list(program.keys()):
        # print(list(program[prog].keys()))
        print(f"**** {program[prog]['programId']} ****", file=fid)

        # Excel output?
        ws = None
        if xls is not None:
            wb = Workbook()
            ws = wb.active

        # First pass to count and record groups, also show the ATOM note
        for item in list(program[prog].keys()):
            if 'INFO' in item:
                if 'ATOM' in program[prog][item]['title']:
                    print(f"{program[prog][item]['title']}: {program[prog][item]['text']}\n", file=fid)
                    # Comment
                    if xls is not None and ws is not None:
                        ws['A1'] = program[prog]['programId']
                        ws['A2'] = 'ATOMS'
                        ws['B2'] = program[prog][item]['text']
            if 'GROUP' in item:
                # print(item, program[prog][item]['name'])
                # print(program[prog][item])
                grpnum.append(int(item.split('-')[1]))
                grplist.append(item.split('-')[0])

        if len(grpnum) > 0:
            # Second pass to put the groups in the same order as in the program
            # print(grplist)
            # print(grpnum)
            isrt = np.argsort(grpnum)
            # print(grpnum)
            for ii in isrt:
                group = str(grplist[ii]) + '-' + str(grpnum[ii])
                print(group, program[prog][group]['name'], file=fid)
                group_proc(program[prog][group], sel_obs_class=sel_obs_class, sel_obs_status=sel_obs_status,
                           fid=fid, wb=wb)

        # Process any observations at the root level
        group_proc(program[prog], sel_obs_class=sel_obs_class, sel_obs_status=sel_obs_status, fid=fid, wb=wb)
        print('', file=fid)

    if xls is not None:
        wb.save(filename=xls)


def printseq(sequence, comment: str = '', csv: bool = False, path: str = ''):
    """
    Print basic configuration information about a sequence, with an option to write to a csv file.
    """
    atom = '1'
    out_file = None
    if csv and path:
        obsid = sequence[0]['ocs:observationId']
        filename = os.path.join(path, obsid + '_seq.csv')
        out_file = open(filename, 'w')
        print(f'comment,{comment}', file=out_file)
        print('datalab,class,exptime,coadds,inst,fpu,filter_name,disperser,wavelength,p,q,atom', file=out_file)

    for step in list(sequence):
        datalab = step['observe:dataLabel']
        observe_class = step['observe:class']
        exptime = step['observe:exposureTime']
        inst = step['instrument:instrument']
        # print(inst, fpuinst[inst])
        fpu = step[fpuinst[inst]]
        filter_name = step["instrument:filter_name"] if 'instrument:filter_name' in step.keys() else 'None'
        wavelength = step['instrument:observingWavelength']
        if 'GMOS' in inst:
            coadds = '1'
            # convert wavelength to microns
            # wavelength = '{:5.3f}'.format(float(wavelength) / 1000.)
        else:
            coadds = step['observe:coadds']
        disperser = step['instrument:disperser']
        p = step['telescope:p'] if 'telescope:p' in step.keys() else '0.0'
        q = step['telescope:q'] if 'telescope:q' in step.keys() else '0.0'

        print(f'{datalab:25} {observe_class:10} {exptime:7} {coadds:3} {inst:10} {fpu:20} {filter_name:12} '
              f'{disperser:12} {wavelength:7} {p:5} {q:5}')
        if out_file is not None:
            print(f'{datalab},{observe_class},{exptime},{coadds},{inst},{fpu},{filter_name},{disperser},{wavelength},'
                  f'{p},{q},{atom}', file=out_file)

    if out_file is not None:
        out_file.close()


def seqxlsx(sequence, comment: str = '', path: str = '') -> NoReturn:
    """
    Write sequence information to an Excel spreadsheet.
    """
    obsid = sequence[0]['ocs:observationId']
    filename = os.path.join(path, obsid + '_seq.xlsx')
    wb = Workbook()
    ws = wb.active

    atom = '1'

    # Comment
    ws['A1'] = 'comment'
    ws['B1'] = comment

    # Columns
    columns = ['datalab', 'class', 'inst', 'exptime', 'coadds', 'fpu', 'filter_name',
               'disperser', 'wavelength', 'p', 'q', 'atom']

    row = 2
    for col_idx, col_data in enumerate(columns):
        ws.cell(column=col_idx + 1, row=row, value=f'col_data')
    row += 1

    for step in list(sequence):
        inst = step['instrument:instrument']
        # print(inst, fpuinst[inst])

        if 'GMOS' in inst:
            coadds = '1'
            # convert wavelength to microns
            # wavelength = '{:5.3f}'.format(float(wavelength) / 1000.)
        else:
            coadds = step['observe:coadds']

        filter_name = step["instrument:filter_name"] if 'instrument:filter_name' in step.keys() else 'None'

        p = step['telescope:p'] if 'telescope:p' in step.keys() else '0.0'
        q = step['telescope:q'] if 'telescope:q' in step.keys() else '0.0'

        data = [step['observe:dataLabel'], step['observe:class'], inst, float(step['observe:exposureTime']),
                int(coadds), step[fpuinst[inst]], filter_name, step['instrument:disperser'],
                float(step['instrument:observingWavelength']), float(p), float(q), int(atom)]
        print(data)

        for col_idx, col_data in enumerate(data):
            ws.cell(column=col_idx + 1, row=row, value=f'{col_data}')
        row += 1

    wb.save(filename)


def readseq(file: str, path: str):
    """
    Read sequence information from a csv file.
    """
    sequence = {}

    with open(os.path.join(path, file), 'r') as csv_file:
        # Read and parse csv file: first line is a comment, second has column headings
        nline = 0
        for line in csv_file:
            values = line.rstrip('\n').split(',')
            if nline == 0:
                sequence['comment'] = values[1]
            elif nline == 1:
                columns = list(values)
                print(columns)
                for col in columns:
                    sequence[col.strip(' ')] = []
            else:
                for i, val in enumerate(values):
                    sequence[columns[i].strip(' ')].append(val.strip(' '))
            nline += 1

    return sequence


def xlsxseq(file: str, path: str):
    """
    Read sequence information from an Excel spreadsheet.
    """
    sequence = {}

    wb = load_workbook(filename=os.path.join(path, file))
    ws = wb.active

    row = 1
    sequence['comment'] = ws.cell(column=2, row=row).value
    row += 1

    columns = []
    # Eventually ready the number of columns in the sheet
    for idx in range(26):
        col = ws.cell(column=idx + 1, row=row).value
        if col is not None:
            columns.append(col)
            sequence[col] = []
        else:
            break
    row += 1

    while ws.cell(column=1, row=row).value is not None:
        for col_idx, col in enumerate(columns):
            sequence[col].append(ws.cell(column=col_idx + 1, row=row).value)
        row += 1

    return sequence


def xlsxatoms(file: str, path: str, sheet: str = 'None', verbose=False):
    # Read a spreadsheet created by findatoms
    wb = load_workbook(os.path.join(path, file))

    atoms_dict = {}

    sheets = []
    if sheet == 'None':
        # Read all sheets except the first
        sheets = wb.sheetnames
        sheets.remove('Sheet')
    elif sheet in wb.sheetnames:
        sheets = [sheet]
    else:
        print(f"Sheet {sheet} not found.")

    for sheet in sheets:
        ws = wb[sheet]
        print(f"Sheet {ws.title}")
        sequence = {}
        columns = []

        # Eventually read the number of columns in the sheet
        row = 1
        for idx in range(16):
            col = ws.cell(column=idx + 1, row=row).value
            if col is not None:
                columns.append(col)
                sequence[col] = []
            else:
                break
        row += 1

        natom = 0
        atoms = []
        classes = []
        guiding = []
        qastates = []
        obs_id = None

        while ws.cell(column=1, row=row).value is not None:
            nextatom = False
            for idx, col in enumerate(columns):
                sequence[col].append(ws.cell(column=idx + 1, row=row).value)
                # print(sequence['datalab'][-1], sequence['inst'][-1], sequence['atom'][-1])

            if natom == 0:
                datalab = sequence['datalab'][-1]
                obs_id = datalab[0:datalab.rfind('-')]

            if natom == 0 or (natom > 0 and sequence['atom'][-1] != sequence['atom'][-2]):
                nextatom = True

            # New atom?
            if nextatom:
                # Get class, qastate, guiding for previous atom
                if natom > 0:
                    atoms[-1]['qa_state'] = select_qastate(qastates)
                    if atoms[-1]['qa_state'] != 'NONE':
                        atoms[-1]['observed'] = True
                    atoms[-1]['class'] = select_obsclass(classes)
                    if verbose:
                        print('QA states: ', qastates)
                        print('Classes: ', classes)
                        print('Guiding: ', guiding)
                    atoms[-1]['guide_state'] = any(guiding)
                    atoms[-1]['wavelength'] = sequence['wavelength'][-1]
                    atoms[-1]['required_resources']['inst'] = sequence['inst'][-1]
                    atoms[-1]['required_resources']['filter'] = sequence['filter'][-1]
                    atoms[-1]['required_resources']['disperser'] = sequence['disperser'][-1]
                    atoms[-1]['required_resources']['fpu'] = sequence['fpu'][-1]

                    # Print basic atom info
                    print(" \t exec_time: {:7.2f}, prog_time: {:7.2f}, part_time: {:7.2f}, guide_state: {}".
                          format(atoms[-1]['exec_time'], atoms[-1]['prog_time'], atoms[-1]['part_time'],
                                 atoms[-1]['guide_state']))
                natom += 1
                atoms.append({'id': natom, 'exec_time': 0.0, 'prog_time': 0.0, 'part_time': 0.0,
                              'class': 'NONE', 'observed': False, 'qa_state': 'NONE', 'guide_state': False,
                              'wavelength': 0.0,
                              'required_resources': {'inst': 'NONE', 'filter': 'NONE', 'disperser': 'NONE',
                                                     'fpu': 'NONE'}})
                classes = []
                guiding = []
                qastates = []

            qastates.append(sequence['qa_state'][-1])
            classes.append(sequence['class'][-1])
            guiding.append(sequence['guiding'][-1])

            atoms[-1]['exec_time'] += sequence['exec_time'][-1]

            atomlabel = natom
            if 'PARTNERCAL' in sequence['class'][-1]:
                atomlabel *= 10
                atoms[-1]['part_time'] += sequence['exec_time'][-1]
            else:
                atoms[-1]['prog_time'] += sequence['exec_time'][-1]

            row += 1

        # Get class/state for last atom
        if natom > 0:
            atoms[-1]['qa_state'] = select_qastate(qastates)
            if atoms[-1]['qa_state'] != 'NONE':
                atoms[-1]['observed'] = True
            atoms[-1]['class'] = select_obsclass(classes)
            if verbose:
                print('QA states: ', qastates)
                print('Classes: ', classes)
                print('Guiding: ', guiding)
            atoms[-1]['guide_state'] = any(guiding)
            atoms[-1]['wavelength'] = sequence['wavelength'][-1]
            atoms[-1]['required_resources']['inst'] = sequence['inst'][-1]
            atoms[-1]['required_resources']['filter'] = sequence['filter'][-1]
            atoms[-1]['required_resources']['disperser'] = sequence['disperser'][-1]
            atoms[-1]['required_resources']['fpu'] = sequence['fpu'][-1]

            # Print basic atom info
            print(f' \t exec_time: {atoms[-1]["exec_time"]:7.2f}, prog_time: {atoms[-1]["prog_time"]:7.2f}, '
                  f'part_time: {atoms[-1]["part_time"]:7.2f}, guide_state: {atoms[-1]["guide_state"]}')

        if obs_id is not None:
            atoms_dict[obs_id] = atoms

    wb.close()
    return atoms_dict


if __name__ == '__main__':
    path = os.path.join(ROOT_DIR, 'app', 'data')
    print(path)

    # programs = ['GN-2018B-Q-101', 'GN-2018B-Q-106', 'GN-2018B-FT-206', ]
    # programs = ['GN-2018B-DD-104', 'GS-2018B-Q-226']
    programs = ['GS-2018B-Q-102']
    f = sys.stdout
    # f = open(os.path.join(path, 'atoms_test.txt'), 'w')
    for progid in programs:
        xlsout = os.path.join(path, progid + '.xlsx')
        program = odb_json(progid, path=path)
        prog_proc(program,
                  sel_obs_class=frozenset(['SCIENCE', 'PROGCAL', 'PARTNERCAL']),
                  sel_obs_status=frozenset(['READY', 'ONGOING', 'OBSERVED']),
                  xls=xlsout, fid=f)
    if f != sys.stdout:
        f.close()

    atoms = xlsxatoms('GS-2018B-Q-102.xlsx', path, '23')
    print(atoms)
