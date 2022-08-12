import csv
import logging
import os
from datetime import datetime, timedelta
from typing import Dict, NoReturn

from openpyxl import load_workbook

from common.helpers import str_to_bool
from common.minimodel.site import Site


class ResourceMock:
    def __init__(self, path: str):
        self.path = os.path.join(os.getcwd(), path)
        self.fpu = {}
        self.fpur = {}
        self.grat = {}
        self.fpu_to_barcode = {}
        self.instruments = {Site.GS: {}, Site.GN: {}}
        self.mode = {Site.GS: {}, Site.GN: {}}
        self.lgs = {Site.GS: {}, Site.GN: {}}
        self.ifu = {Site.GS: {}, Site.GN: {}}

    def _load_fpu(self, name: str, site: Site) -> tuple[dict[datetime, str], dict[datetime, list[str]]]:
        ifu = {}
        barcodes = {}
        with open(os.path.join(self.path, f'GMOS{site.name}_{name}201789.txt')) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ifu[datetime.strptime(row[0].strip(), "%Y-%m-%d")] = row[1].strip()
                barcodes[datetime.strptime(row[0].strip(), "%Y-%m-%d")] = [i.strip() for i in row[2:]]
        return ifu, barcodes

    def _load_gratings(self, site: Site) -> dict[datetime, list[str]]:
        out_dict = {}
        with open(os.path.join(self.path, f'GMOS{site.name}_GRAT201789.txt')) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                out_dict[datetime.strptime(row[0].strip(), "%Y-%m-%d")] = [i.strip().replace('+', '') for i in row[1:]]
                out_dict[datetime.strptime(row[0].strip(), "%Y-%m-%d")].append('MIRROR')
        return out_dict

    def _load_fpu_to_barcodes(self, site: Site) -> Dict[str, str]:
        out_dict = {}
        with open(os.path.join(self.path, f'gmos{site.name}_fpu_barcode.txt')) as f:
            for row in f:
                fpu, barcode = row.split()
                out_dict[fpu] = barcode
        return out_dict

    @staticmethod
    def _previous(items, pivot):
        # Return date equal or previous to pivot
        tdmin = timedelta.min
        tdzero = timedelta(days=0)
        result = None

        for item in items:
            diff = item - pivot
            if tdzero >= diff > tdmin:
                result = item
                tdmin = diff
        return result

    def _excel_reader(self) -> NoReturn:
        workbook = load_workbook(filename=os.path.join(self.path, '2018B-2019A Telescope Schedules.xlsx'))
        for site in Site:
            sheet = workbook[site.name]
            for row in sheet.iter_rows(min_row=2):
                date = row[0].value
                self.instruments[site][date] = ['Flamingos2' if c.value == 'F2' else c.value for c in row[3:]]
                self.mode[site][date] = row[1].value
                self.lgs[site][date] = str_to_bool(row[2].value)

        if not self.instruments or not self.mode or not self.lgs:
            raise Exception("Problems reading spreadsheet..")

    def connect(self) -> NoReturn:
        """
        Allows the mock to load all the data locally, emulating a connection to the API.
        """
        print('Get Resource data...')
        for site in Site:
            self.ifu[site]['FPU'], self.fpu[site] = self._load_fpu('FPU', site)
            self.ifu[site]['FPUr'], self.fpur[site] = self._load_fpu('FPUr', site)
            self.grat[site] = self._load_gratings(site)
            self.fpu_to_barcode[site] = self._load_fpu_to_barcodes(site)

        if not self.fpu or not self.fpur or not self.grat:
            raise Exception("Problems on reading files...")

        self._excel_reader()

    def _get_info(self, info: str, site: Site, date_str: str):
        date = datetime.strptime(date_str, "%Y-%m-%d")

        info_types = {'fpu': self.fpu[site],
                      'fpur': self.fpur[site],
                      'grat': self.grat[site],
                      'instr': self.instruments[site],
                      'LGS': self.lgs[site],
                      'mode': self.mode[site],
                      'fpu-ifu': self.ifu[site]['FPU'],
                      'fpur-ifu': self.ifu[site]['FPUr']}

        if info in info_types:
            previous_date = ResourceMock._previous(info_types[info].keys(), date)
            return info_types[info][previous_date]

        else:
            logging.warning(f'No information about {info} is stored')
            return None

    def get_night_resources(self, sites, night_date):
        def night_info(info_name: str):
            return {site: self._get_info(info_name, site, night_date) for site in sites}

        fpu = night_info('fpu')
        fpur = night_info('fpur')
        gratings = night_info('grat')
        instruments = night_info('instr')
        lgs = night_info('LGS')
        modes = night_info('mode')
        ifus = {'FPU': night_info('fpu-ifu'), 'FPUr': night_info('fpur-ifu')}
        return Resources(fpu, fpur, gratings, instruments, lgs, modes, self.fpu_to_barcode, ifus)
