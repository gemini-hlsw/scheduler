#!/usr/bin/env python
# coding: utf-8

# Code for manipulating ODB Extractor json files and initial sequence atom definitions
# Bryan Miller
# 2021-11-24

import os
import json
import gzip
import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import find_peaks
import requests

from openpyxl import Workbook
from openpyxl import load_workbook


fpuinst = {'GSAOI': 'instrument:utilityWheel', 'GPI': 'instrument:observingMode', 'Flamingos2': 'instrument:fpu',
           'NIFS': 'instrument:mask', 'GNIRS': 'instrument:slitWidth', 'GMOS-N': 'instrument:fpu',
           'GMOS-S': 'instrument:fpu', 'NIRI': 'instrument:mask'}

gpi_filter_wav = {'Y': 1.05, 'J': 1.25, 'H': 1.65, 'K1': 2.05, 'K2': 2.25}
nifs_filter_wav = {'ZJ': 1.05, 'JH': 1.25, 'HK': 2.20}


def find_filter(input, filter_dict):
    """Match input string with filter list (in dictionary)"""

    filter = ''
    filters = list(filter_dict.keys())
    for filt in filters:
        if filt in input:
            filter = filt
            break
    return filter


def uniquelist(seq):
    # Make a list of unique values
    # http://stackoverflow.com/questions/480214/how-do-you-remove-duplicates-from-a-list-in-python-whilst-preserving-order
    seen = set()
    seen_add = seen.add
    return [x for x in seq if not (x in seen or seen_add(x))]


def odb_json(progid, verbose=False):
    """
    Download json of ODB program information

    Parameters
        progid:  Program ID of program to extract

    Return
        request_json:   JSON query result as a list of dictionaries
    """

    if progid == "":
        print('odb_json: program id not given.')
        raise ValueError('Program id not given.')

    response = requests.get(
        'http://gnodbscheduler.hi.gemini.edu:8442/programexport?id=' + progid)
    try:
        response.raise_for_status()
    except requests.exceptions.HTTPError as exc:
        print('odb_json: request failed: {}'.format(response.text))
        raise exc
    else:
        request_json = response.json()

    if verbose:
        print(response.url)
    # print(response.text)

    return request_json

# --------------


def guide_state(step):
    """Determine if guiding is on/off for a sequence step"""
    # One could also extract the guider used if needed
    guiding = False
    for key in list(step.keys()):
        if 'guideWith' in key:
            if step[key] == 'guide':
                guiding = True
                break
    return guiding

# --------------


def select_qastate(states):
    """Return the qastate based on precedence

        states: list of observe states from the ODB extractor obsLog
    """
    qastate = ''

    # Precedence order for observation classes.
    qastate_order = ['NONE', 'UNDEFINED', 'FAIL', 'USABLE', 'PASS']

    # Set the qastate for the entire observation based on precedence
    for state in qastate_order:
        if state in states:
            qastate = state
            break

    return qastate

# --------------


def select_obsclass(classes):
    """Return the obsclass based on precedence

        classes: list of observe classes from the ODB extractor
    """
    obsclass = ''

    # Precedence order for observation classes.
    obsclass_order = ['SCIENCE', 'PROGCAL', 'PARTNERCAL', 'ACQ', 'ACQCAL', 'DAYCAL']

    # Set the obsclass for the entire observation based on obsclass precedence
    for oclass in obsclass_order:
        if oclass in classes:
            obsclass = oclass
            break

    return obsclass

# --------------


def autocorr_lag(x, plot=False):
    '''Test for patterns with auto-correlation'''

    # Auto correlation
    result = np.correlate(x, x, mode='full')
    corrmax = np.max(result)
    if corrmax != 0.0:
        result = result / corrmax
    #     print(result)
    #     print(result.size, result.size//2)
    if plot:
        plt.plot(result[result.size // 2:])

    # Pattern offset using first prominent peak
    # lag = np.argmax(result[result.size//2 + 1:]) + 1
    peaks, prop = find_peaks(result[result.size // 2:], height=(0, None), prominence=(0.25, None))
    #     print(peaks)
    #     print(prop)
    lag = 0
    if len(peaks) > 0:
        lag = peaks[0]

    return lag

# --------------


def findatoms(observation):
    """Analyze a json observing sequence from the ODB and define atoms."""

    classes = []
    guiding = []
    qastates = []
    atoms = []
    natom = 0

    # Make dictionary out of obslog to get QA state
    obslog = {}
    datalabels = []
    if 'obsLog' in observation.keys():
        for log_entry in observation['obsLog']:
            obslog[log_entry['label']] = {'qaState': log_entry['qaState'], 'filename': log_entry['filename']}
        datalabels = list(obslog.keys())

    # Sequence analysis
    sequence = observation['sequence']
    nsteps = len(sequence)

    # First pass to analyze offsets and exptimes/coadds
    exptimes = []
    coadds_list = []
    qoffsets = []
    poffsets = []
    for ii, step in enumerate(sequence):
        # just pick on-sky exposures
        if step['observe:observeType'].upper() not in ['FLAT', 'ARC', 'DARK', 'BIAS']:
            step_keys = list(step.keys())

            if 'telescope:p' in step_keys:
                p = float(step['telescope:p'])
            else:
                p = 0.0
            if 'telescope:q' in step_keys:
                q = float(step['telescope:q'])
            else:
                q = 0.0
            if 'observe:coadds' in step_keys:
                coadds = int(step['observe:coadds'])
            else:
                coadds = 1
            poffsets.append(p)
            qoffsets.append(q)
            exptimes.append(float(step['observe:exposureTime']))
            coadds_list.append(coadds)

    #     print('poffsets: ', poffsets)
    #     print('qoffsets: ', qoffsets)
    # Analyze patterns using auto-correlation
    # The lag is the length of any pattern, 0 means no repeating pattern
    plag = 0
    qlag = 0
    if len(poffsets) > 1:
        plag = autocorr_lag(np.array(poffsets))
    if len(qoffsets) > 1:
        qlag = autocorr_lag(np.array(qoffsets))

    # Special cases
    if plag == 0 and qlag == 0 and len(qoffsets) == 4:
        # single ABBA pattern, which the auto-correlation won't find
        if qoffsets[0] == qoffsets[3] and qoffsets[1] == qoffsets[2]:
            qlag = 4
    elif len(qoffsets) == 2:
        # If only two steps, put them together, might be AB, also silly to split only two steps
        qlag = 2

    offset_lag = qlag
    if plag > 0 and plag != qlag:
        offset_lag = 0
    print('lags: ', plag, qlag, offset_lag)

    # Changes in exptimes/coadds?
    if len(uniquelist(exptimes)) > 1 or len(uniquelist(coadds_list)) > 1:
        exptime_groups = True
    else:
        exptime_groups = False

    # Second pass to determine atom properties
    npattern = offset_lag
    wavelength_prev = 0.0
    exptime_prev = 0.0
    coadds_prev = 0
    for ii, step in enumerate(sequence):
        nextatom = False
        #         qoffsets = False

        step_keys = list(step.keys())

        datalab = step['observe:dataLabel']
        if datalab in datalabels:
            qastate = obslog[datalab]['qaState']
        else:
            qastate = 'NONE'
        qastates.append(qastate.upper())

        observe_class = step['observe:class']
        classes.append(observe_class.upper())

        guiding.append(guide_state(step))

        exptime = float(step['observe:exposureTime'])
        step_time = step['totalTime'] / 1000.

        inst = step['instrument:instrument']
        #     print(inst, fpuinst[inst])
        if inst == 'Visitor Instrument':
            fpu = 'Visitor'
        else:
            fpu = step[fpuinst[inst]]

        if 'instrument:disperser' in step_keys:
            disperser = step['instrument:disperser']
        elif inst == 'Visitor Instrument':
            disperser = 'Visitor'
        else:
            disperser = 'None'

        if 'instrument:filter' in step_keys:
            filter = step["instrument:filter"]
        elif inst == 'GPI':
            filter = find_filter(fpu, gpi_filter_wav)
        elif inst == 'Visitor Instrument':
            filter = 'Visitor'
        else:
            filter = 'None'

        if inst == 'NIFS' and 'Same as Disperser' in filter:
            for filt in list(nifs_filter_wav.keys()):
                if disperser[0] in filt:
                    filter = filt
                    break

        if inst == 'GPI':
            wavelength = gpi_filter_wav[filter]
        else:
            wavelength = float(step['instrument:observingWavelength'])

        if 'observe:coadds' in step_keys:
            coadds = int(step['observe:coadds'])
        else:
            coadds = 1

        if 'telescope:p' in step_keys:
            p = float(step['telescope:p'])
        else:
            p = 0.0
        if 'telescope:q' in step_keys:
            q = float(step['telescope:q'])
        #             qoffsets = True
        else:
            q = 0.0

            # Any wavelength/filter change is a new atom
        if wavelength != wavelength_prev:
            nextatom = True
            print('Atom for wavelength')

        # A change in exposure time or coadds is a new atom
        if step['observe:observeType'].upper() not in ['FLAT', 'ARC', 'DARK', 'BIAS'] \
                and exptime != exptime_prev or coadds != coadds_prev:
            nextatom = True
            print('Atom for exposure time change')

        # Offsets - a new offset pattern is a new atom
        print('npattern: ', npattern)
        if step['observe:observeType'].upper() not in ['FLAT', 'ARC', 'DARK', 'BIAS'] \
                and not (offset_lag == 0 and exptime_groups == True):
            npattern -= 1
            if npattern < 0:
                nextatom = True
                print('Atom from offset pattern')
                npattern = offset_lag - 1

        if nextatom:
            # Get class, qastate, guiding for previous atom
            if natom > 0:
                print(qastates)
                atoms[-1]['qa_state'] = select_qastate(qastates)
                if atoms[-1]['qa_state'] != 'NONE':
                    atoms[-1]['observed'] = True
                print(classes)
                atoms[-1]['class'] = select_obsclass(classes)
                print(guiding)
                atoms[-1]['guide_state'] = any(guiding)
                atoms[-1]['wavelength'] = wavelength
                atoms[-1]['required_resources']['inst'] = inst
                atoms[-1]['required_resources']['filter'] = filter
                atoms[-1]['required_resources']['disperser'] = disperser
                atoms[-1]['required_resources']['fpu'] = fpu

                # New atom
            natom += 1
            atoms.append({'id': natom, 'exec_time': 0.0, 'prog_time': 0.0, 'part_time': 0.0,
                          'class': 'NONE', 'observed': False, 'qa_state': 'NONE', 'guide_state': False,
                          'wavelength': 0.0,
                          'required_resources': {'inst': 'NONE', 'filter': 'NONE', 'disperser': 'NONE',
                                                 'fpu': 'NONE'}})
            classes = []
            guiding = []
            if step['observe:observeType'].upper() in ['FLAT', 'ARC', 'DARK', 'BIAS'] \
                    and npattern == 0:
                npattern = offset_lag

        atoms[-1]['exec_time'] += step_time

        atomlabel = natom
        if 'partnerCal' in observe_class:
            atomlabel *= 10
            atoms[-1]['part_time'] += step_time
        else:
            atoms[-1]['prog_time'] += step_time

        print('{:22} {:12} {:7.2f} {:3d} {:10} {:15} {:12} {:12} {:7.4f} {:5.2f} {:5.2f} {:3d}'.format(datalab,
                                                                                                       observe_class,
                                                                                                       exptime, coadds,
                                                                                                       inst, fpu,
                                                                                                       filter,
                                                                                                       disperser,
                                                                                                       wavelength, p, q,
                                                                                                       atomlabel))

        wavelength_prev = wavelength
        if step['observe:observeType'].upper() not in ['FLAT', 'ARC', 'DARK', 'BIAS']:
            exptime_prev = exptime
            coadds_prev = coadds

    #     print(atoms)
    # Get class/state for last atom
    if natom > 0:
        print(qastates)
        atoms[-1]['qa_state'] = select_qastate(qastates)
        if atoms[-1]['qa_state'] != 'NONE':
            atoms[-1]['observed'] = True
        print(classes)
        atoms[-1]['class'] = select_obsclass(classes)
        print(guiding)
        atoms[-1]['guide_state'] = any(guiding)
        atoms[-1]['wavelength'] = wavelength
        atoms[-1]['required_resources']['inst'] = inst
        atoms[-1]['required_resources']['filter'] = filter
        atoms[-1]['required_resources']['disperser'] = disperser
        atoms[-1]['required_resources']['fpu'] = fpu

    return atoms


def group_proc(group, sel_obs_class=['SCIENCE', 'PROGCAL', 'PARTNERCAL', 'ACQ', 'ACQCAL', 'DAYCAL'],
               sel_obs_status=['PHASE_2', 'FOR_REVIEW', 'IN_REVIEW', 'FOR_ACTIVATION', 'ON_HOLD', 'READY',
                               'ONGOING', 'OBSERVED', 'INACTIVE']):
    """Process observations within groups"""

    obsnum = []
    for item in list(group.keys()):
        obsid = ''
        if 'OBSERVATION' in item:
            #         obsid = program[prog][group][item]['sequence'][0]['ocs:observationId']
            obsid = group[item]['observationId']
            obsnum.append(int(item.split('-')[1]))
    #             print(f" \t {item, obsnum[-1], obsid}")
    #         else:
    #             print(item, group[item])

    if len(obsnum) > 0:
        isrt = np.argsort(obsnum)
        for ii in isrt:
            obs_program_used = 0.0
            obs_partner_used = 0.0
            item = 'OBSERVATION_BASIC-' + str(obsnum[ii])
            #     obsid = program[prog][group][item]['sequence'][0]['ocs:observationId']
            obsid = group[item]['observationId']
            print(f" \t {obsnum[ii], obsid}")
            obs_class = group[item]['obsClass'].upper()
            phase2stat = group[item]['phase2Status'].upper()
            obs_stat = group[item]['obsStatus'].upper()
            print(obs_class, phase2stat, obs_stat)
            if obs_class in sel_obs_class and obs_stat in sel_obs_status:
                # Atoms in each sequence
                atoms = findatoms(group[item])
                # Summary of atoms
                classes = []
                qastates = []
                for atom in atoms:
                    print('Atom ', atom['id'])
                    for key in atom.keys():
                        print(f" \t {key}: {atom[key]}")
                        if key == 'class':
                            classes.append(atom[key])
                        if key == 'qa_state':
                            qastates.append(atom[key])
                            if atom[key].upper() == 'PASS':
                                obs_program_used += atom['prog_time']
                                obs_partner_used += atom['part_time']
                obsclass = select_obsclass(classes)
                print(f"Obsclass: {obsclass}")
                obs_qastate = select_qastate(qastates)
                print(f"QAstate (atoms): {obs_qastate}")
                print(f"qaState (ODB): {group[item]['qaState']}")
                if group[item]['qaState'].upper() == 'PASS':
                    if group[item]['obsClass'] in ['science', 'progCal']:
                        obs_program_used += float(group[item]['setupTime']) / 1000.
                    elif group[item]['obsClass'] in ['partnerCal']:
                        obs_partner_used += float(group[item]['setupTime']) / 1000.

                print(f"program_used: {obs_program_used}")
                print(f"partner_used: {obs_partner_used}")

            print()

    return


def prog_proc(program, sel_obs_class=['SCIENCE', 'PROGCAL', 'PARTNERCAL', 'ACQ', 'ACQCAL', 'DAYCAL'],
              sel_obs_status=['PHASE_2', 'FOR_REVIEW', 'IN_REVIEW', 'FOR_ACTIVATION', 'ON_HOLD', 'READY',
                              'ONGOING', 'OBSERVED', 'INACTIVE']):
    """Process top-level of program"""

    grpnum = []
    grplist = []
    for prog in list(program.keys()):
        #     print(list(program[prog].keys()))

        # Any observations at the root level?
        group_proc(program[prog], sel_obs_class=sel_obs_class, sel_obs_status=sel_obs_status)

        # First pass to count and record groups
        for item in list(program[prog].keys()):
            if 'GROUP' in item:
                #                 print(item, program[prog][item]['name'])
                #         print(program[prog][item])
                grpnum.append(int(item.split('-')[1]))
                grplist.append(item.split('-')[0])

        #         gmax = np.max(grpnum)

        if len(grpnum) > 0:
            # Second pass to put the groups in the same order as in the program
            print(grplist)
            print(grpnum)
            isrt = np.argsort(grpnum)
            #     print(grpnum)
            for ii in isrt:
                group = grplist[ii] + '-' + str(grpnum[ii])
                print(group, program[prog][group]['name'])
                group_proc(program[prog][group], sel_obs_class=sel_obs_class, sel_obs_status=sel_obs_status)

    return


def printseq(sequence, comment='', csv=False, path=''):
    '''Print basic configuration information about a sequence, with an option to write to a csv file'''

    atom = '1'
    if csv and path != '':
        obsid = sequence[0]['ocs:observationId']
        filename = os.path.join(path, obsid + '_seq.csv')
        f = open(filename, 'w')
        print('{},{}'.format('comment', comment), file=f)
        print('{},{},{},{},{},{},{},{},{},{},{},{}'.format('datalab', 'class', 'exptime', 'coadds', 'inst', 'fpu',
                                                        'filter_name', 'disperser', 'wavelength', 'p', 'q', 'atom'), file=f)

    for step in list(sequence):
        datalab = step['observe:dataLabel']
        observe_class = step['observe:class']
        exptime = step['observe:exposureTime']
        inst = step['instrument:instrument']
    #     print(inst, fpuinst[inst])
        fpu = step[fpuinst[inst]]
        if 'instrument:filter_name' in step.keys():
            filter_name = step["instrument:filter_name"]
        else:
            filter_name = 'None'
        wavelength = step['instrument:observingWavelength']
        if 'GMOS' in inst:
            coadds = '1'
            # convert wavelength to microns
#             wavelength = '{:5.3f}'.format(float(wavelength) / 1000.)
        else:
            coadds = step['observe:coadds']
        disperser = step['instrument:disperser']
        if 'telescope:p' in step.keys():
            p = step['telescope:p']
        else:
            p = '0.0'
        if 'telescope:q' in step.keys():
            q = step['telescope:q']
        else:
            q = '0.0'    
        print('{:25} {:10} {:7} {:3} {:10} {:20} {:12} {:12} {:7} {:5} {:5}'.format(datalab, observe_class, exptime, coadds,
                                                                       inst, fpu, filter_name, disperser, wavelength, p, q))
        if csv and path != '':
            print('{},{},{},{},{},{},{},{},{},{},{},{}'.format(datalab, observe_class, exptime, coadds, inst, fpu,
                                                               filter_name, disperser, wavelength, p, q, atom), file=f)

    if csv and path != '':
        f.close()


def seqxlsx(sequence, comment='', path=''):
    '''Write sequence information to an Excel spreadsheet'''

    obsid = sequence[0]['ocs:observationId']
    filename = os.path.join(path, obsid + '_seq.xlsx')
    wb = Workbook()
    ws = wb.active
    
    atom = '1'
    
    # Comment
    ws['A1'] = 'comment'
    ws['B1'] = comment

    # Columns
    columns = ['datalab', 'class', 'inst', 'exptime', 'coadds', 'fpu', 'filter_name',
               'disperser', 'wavelength', 'p', 'q', 'atom']
    
    row = 2
    for ii, col in enumerate(columns):
        _ = ws.cell(column=ii+1, row=row, value="{0}".format(col))
    row += 1
        
#     print('{},{}'.format('comment', comment), file=f)
#     print('{},{},{},{},{},{},{},{},{},{},{}'.format('datalab', 'class', 'exptime', 'coadds', 'inst', 'fpu', 
#                                                            'disperser', 'wavelength', 'p', 'q', 'atom'), file=f)

    for step in list(sequence):
        data = []
        data.append(step['observe:dataLabel'])
        data.append(step['observe:class'])
        inst = step['instrument:instrument']
        data.append(inst)
    #     print(inst, fpuinst[inst])
        data.append(float(step['observe:exposureTime']))
        if 'GMOS' in inst:
            coadds = '1'
            # convert wavelength to microns
#             wavelength = '{:5.3f}'.format(float(wavelength) / 1000.)
        else:
            coadds = step['observe:coadds']
        data.append(int(coadds))
        data.append(step[fpuinst[inst]])
        if 'instrument:filter_name' in step.keys():
            filter_name = step["instrument:filter_name"]
        else:
            filter_name = 'None'
        data.append(filter_name)
        data.append(step['instrument:disperser'])
        data.append(float(step['instrument:observingWavelength']))
        if 'telescope:p' in step.keys():
            p = step['telescope:p']
        else:
            p = '0.0'
        data.append(float(p))
        if 'telescope:q' in step.keys():
            q = step['telescope:q']
        else:
            q = '0.0'  
        data.append(float(q))
        data.append(int(atom))
        print(data)
        
        for ii in range(len(columns)):
            _ = ws.cell(column=ii+1, row=row, value=data[ii])
        row += 1
    
    wb.save(filename)
    return


def readseq(file, path):
    '''Read sequence information from a csv file'''

    filename = os.path.join(path, file)
    f = open(filename, 'r')
    
    sequence = {}
    
    # Read and parse csv file: first line is a comment, second has column headings
    nline = 0
    for line in f:
#         line = line.rstrip('\n')
        values = line.rstrip('\n').split(',')
        if nline == 0:
            sequence['comment'] = values[1]
        elif nline == 1:
            columns = list(values)
            print(columns)
            for col in columns:
                sequence[col.strip(' ')] = []
        else:
            for i, val in enumerate(values):
                sequence[columns[i].strip(' ')].append(val.strip(' '))
        nline += 1
        
    f.close()
    
    return sequence


def xlsxseq(file, path):
    '''Read sequence information from an Excel spreadsheet'''

    filename = os.path.join(path, file)
    
    wb = load_workbook(filename=filename)
    ws = wb.active
    
    sequence = {}
    
    row = 1
    sequence['comment'] = ws.cell(column=2, row=row).value
    row += 1
    
    columns = []
    # Eventually ready the number of columns in the sheet
    for ii in range(26):
        col = ws.cell(column=ii+1, row=row).value
        if col is not None:
            columns.append(col)
            sequence[col] = []
        else:
            break
    row += 1
#     print(columns)

    while ws.cell(column=1, row=row).value is not None:
        for jj, col in enumerate(columns):
            sequence[col].append(ws.cell(column=jj+1, row=row).value)
        row += 1
    
    return sequence


if __name__ == '__main__':

    path = './data'
    print(path)

    # Tellurics
    file = 'GN-2018B-Q-101.json.gz'

    # with open(os.path.join(path, file), 'r') as f:
    #     program = json.load(f)
    #     program = f.read()
    with gzip.open(os.path.join(path, file), 'r') as fin:
        json_bytes = fin.read()  # 3. bytes (i.e. UTF-8)

    json_str = json_bytes.decode('utf-8')
    program = json.loads(json_str)

    # print(list(program.keys()))

    # Top-level program information
    print('Evaluating: ', file)
    print('Top-level program information')
    for prog in list(program.keys()):
    #     print(list(program[prog].keys()))
        for item in list(program[prog].keys()):
            print(item)
    #         print(program[prog][item])
            if 'INFO' in item:
                print(f"\t {program[prog][item]['title']}")
            if all(type not in item for type in ['INFO', 'GROUP']):
                print(f" \t {program[prog][item]}")
    print()

    prog = list(program.keys())[0]
    group = 'GROUP_GROUP_SCHEDULING-23'
    print(group)
    obsnum = []
    for item in list(program[prog][group].keys()):
        obsid = ''
        if 'OBSERVATION' in item:
#         obsid = program[prog][group][item]['sequence'][0]['ocs:observationId']
            obsid = program[prog][group][item]['observationId']
            obsnum.append(item.split('-')[1])
            print(item, obsnum[-1], obsid)
        else:
            print(item, program[prog][group][item])
    print()

    # Sort groups to the order defined in the observation
    print('Sorted into defined order:')
    obsnum.sort()
    # print(obsnum)
    for ii in obsnum:
        item = 'OBSERVATION_BASIC-' + ii
    #     obsid = program[prog][group][item]['sequence'][0]['ocs:observationId']
        obsid = program[prog][group][item]['observationId']
        print(ii, obsid)
    print()

    obs = 'OBSERVATION_BASIC-1'
    print(obs)
    for item in list(program[prog][group][obs].keys()):
        print(item)
        if 'INFO' in item:
                print(f"\t {program[prog][group][obs][item]['title']}")
        if any(type in item for type in ['CONDITIONS', 'TARGETENV']):
            for key in list(program[prog][group][obs][item].keys()):
                print(f"\t {key}")
        if all(type not in item for type in ['INFO', 'sequence', 'obsLog']):
                print(f" \t {program[prog][group][obs][item]}")

    tot_seq_time = 0.0
    for step in list(program[prog][group][obs]['sequence']):
        print(step['observe:dataLabel'], step['observe:class'], step['observe:exposureTime'],
              step['observe:coadds'], step['instrument:instrument'], step['instrument:slitWidth'],
              step['instrument:disperser'], step['instrument:observingWavelength'], step['telescope:p'],
              step['telescope:q'], step['totalTime']/(1000.))
        tot_seq_time += step['totalTime']/(1000.)
    print('Acq overhead: ', program[prog][group][obs]['setupTime']/1000.)
    print('Sequence time:', tot_seq_time)
    exec_time = program[prog][group][obs]['setupTime']/1000. + tot_seq_time
    print('Total observation time: ', exec_time, exec_time/3600.)
    print()

    # printseq(program[prog][group][obs]['sequence'])

    print('Sequence Atoms')
    atoms = findatoms(program[prog][group][obs])
    # Summary of atoms
    for atom in atoms:
        print('Atom ', atom['id'])
        for key in atom.keys():
            print(f" \t {key}: {atom[key]}")
    print()

    targenv = 'TELESCOPE_TARGETENV-2'
    print(targenv)
    for item in list(program[prog][group][obs][targenv].keys()):
        print(item)
        print(program[prog][group][obs][targenv][item])
    print()

    constraints = 'SCHEDULING_CONDITIONS-1'
    print(constraints)
    for item in list(program[prog][group][obs][constraints].keys()):
        print(item, ':', program[prog][group][obs][constraints][item])
    print()
    print()

    print('Process program to define atoms')
    prog_proc(program, sel_obs_class=['SCIENCE', 'PROGCAL', 'PARTNERCAL'],
              sel_obs_status=['READY', 'ONGOING', 'OBSERVED'])
    print()


    # ToO/Timing Window Example
    file = 'GN-2018B-Q-106.json.gz'
    # with open(os.path.join(path, file), 'r') as f:
    #     program = json.load(f)
    with gzip.open(os.path.join(path, file), 'r') as fin:
        json_bytes = fin.read()
    json_str = json_bytes.decode('utf-8')
    program = json.loads(json_str)

    prog_proc(program, sel_obs_class=['SCIENCE', 'PROGCAL', 'PARTNERCAL'],
              sel_obs_status=['READY', 'ONGOING', 'OBSERVED'])
    print()

    # DD/Nonsidereal/Colossus - GMOS-N and NIRI
    file = 'GN-2018B-DD-104.json.gz'
    # with open(os.path.join(path, file), 'r') as f:
    #     program = json.load(f)
    with gzip.open(os.path.join(path, file), 'r') as fin:
        json_bytes = fin.read()  # 3. bytes (i.e. UTF-8)

    json_str = json_bytes.decode('utf-8')
    program = json.loads(json_str)


    # F2
    program = odb_json('GS-2018B-Q-213')
    prog_proc(program, sel_obs_class=['SCIENCE', 'PROGCAL', 'PARTNERCAL'], \
              sel_obs_status=['READY', 'ONGOING', 'OBSERVED'])

    # NIFS
    program = odb_json('GN-2018B-Q-109')
    prog_proc(program, sel_obs_class=['SCIENCE', 'PROGCAL', 'PARTNERCAL'], \
              sel_obs_status=['READY', 'ONGOING', 'OBSERVED'])

    # GSAOI
    program = odb_json('GS-2018B-Q-226')
    prog_proc(program, sel_obs_class=['SCIENCE', 'PROGCAL', 'PARTNERCAL'], \
              sel_obs_status=['READY', 'ONGOING', 'OBSERVED'])

