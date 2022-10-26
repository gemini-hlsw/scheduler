# Copyright (c) 2016-2022 Association of Universities for Research in Astronomy, Inc. (AURA)
# For license information see LICENSE or https://opensource.org/licenses/BSD-3-Clause

from copy import copy
import csv
import os
from datetime import date, datetime, timedelta
from types import MappingProxyType
from typing import Callable, Collection, Dict, Final, FrozenSet, List, NoReturn, Optional, Set, Tuple, final

from lucupy.helpers import str_to_bool
from lucupy.minimodel import ALL_SITES, Resource, Site
from openpyxl import load_workbook

from app.core.meta import Singleton
from definitions import ROOT_DIR


@final
class ResourceMock(metaclass=Singleton):
    """
    This is a mock for the future Resource service, used for OCS.
    It reads data regarding availability of instruments, IFUs, FPUs, MOS masks, etc. at each Site for given dates.

    It can then be queried to receive a set of Resource (usually with barcode IDs, except for instruments) for a
    given site on a given night.

    It caches and reuses Resources by ID as best as possible to minimize the number of Resource objects existing
    at any given time. Since Resource is immutable, this should be fine.

    Note that this is a Singleton class, so new instances do not need to be created.
    """
    # Definition of a day to not have to redeclare constantly.
    _day: Final[timedelta] = timedelta(days=1)

    # These are the converters from the OCS FPU names to the ITCD FPU representations.
    # For example, the ODB query extractor would return:
    #    * 'IFU 2 Slits'
    # which we would want to convert to:
    #    * 'IFU-2'
    # since these are the FPU names used in the GMOS[NS]-FPUr######.txt files.
    _gmosn_ifu_dict = MappingProxyType({
        'IFU 2 Slits': 'IFU-2',
        'IFU Left Slit (blue)': 'IFU-B',
        'IFU Right Slit (red)': 'IFU-R',
        'Longslit 0.25 arcsec': '0.25arcsec',
        'Longslit 0.50 arcsec': '0.5arcsec',
        'Longslit 0.75 arcsec': '0.75arcsec',
        'Longslit 1.00 arcsec': '1.0arcsec',
        'Longslit 1.5 arcsec': '1.5arcsec',
        'Longslit 2.00 arcsec': '2.0arcsec',
        'Longslit 5.00 arcsec': '5.0arcsec',
        'N and S 0.50 arcsec': 'NS0.5arcsec',
        'N and S 0.75 arcsec': 'NS0.75arcsec',
        'N and S 1.00 arcsec': 'NS1.0arcsec',
        'N and S 1.50 arcsec': 'NS1.5arcsec',
        'N and S 2.00 arcsec': 'NS2.0arcsec',
        # TODO: Not in OCS?
        'focus_array_new': 'focus_array_new'
    })

    _gmoss_ifu_dict = MappingProxyType({**_gmosn_ifu_dict, **{
        'IFU N and S 2 Slits': 'IFU-NS-2',
        'IFU N and S Left Slit (blue)': 'IFU-NS-B',
        'IFU N and S Right Slit (red)': 'IFU-NS-R',
        # TODO: Not in OCS?
        'PinholeC': 'PinholeC'
    }})

    def __init__(self, sites: FrozenSet[Site] = ALL_SITES):
        """
        Create and initialize the ResourceMock object with the specified sites.
        """
        self._sites = sites
        self._path = os.path.join(ROOT_DIR, 'mock', 'resource', 'data')

        # This is to avoid recreating repetitive resources.
        # When we first get a resource ID string, create a Resource for it and store it here.
        # Then fetch the Resources from here if they exist, and if they do not, then create a new one as per
        # the _lookup_resource method.
        self._all_resources: Dict[str, Resource] = {}

        # The map from site and date to the set of resources.
        self._resources: Dict[Site, Dict[date, Set[Resource]]] = {site: {} for site in self._sites}

        # Mapping from ITCD FPUs to barcodes. The mapping is site-dependent.
        # The ODB program extractor produces long versions of these names that must be run through the
        # OcsFpuConverter to get the ITCD FPU names.
        self._itcd_fpu_to_barcode: Dict[Site, Dict[str, Resource]] = {site: {} for site in self._sites}

        # Determines whether a night is a part of a laser run.
        self._lgs: Dict[Site, Dict[date, bool]] = {site: {} for site in self._sites}

        for site in self._sites:
            suffix = 's' if site == Site.GS else 'n'
            usuffix = suffix.upper()

            # Load the mappings from the ITCD FPU values to the barcodes.
            self._load_fpu_to_barcodes(site, f'gmos{suffix}_fpu_barcode.txt')

            # Load the FPUrs.
            # This will put both the IFU and the FPU barcodes available on a given date as Resources.
            # Note that for the IFU, we need to convert to a barcode, which is a Resource.
            # This is a bit problematic since we expect a list of strings of Resource IDs, so we have to take its ID.
            self._load_csv(site, f'GMOS{usuffix}_FPUr201789.txt',
                           lambda r: {self._itcd_fpu_to_barcode[site][r[0].strip()].id} | {i.strip() for i in r[1:]})

            # Load the gratings.
            # This will put the mirror and the grating names available on a given date as Resources.
            # TODO: Check Mirror vs. MIRROR. Seems like GMOS uses Mirror.
            self._load_csv(site, f'GMOS{usuffix}_GRAT201789.txt',
                           lambda r: {'Mirror'} | {i.strip().replace('+', '') for i in r})

        # Process the spreadsheet information for instrument, mode, and LGS settings.
        self._load_spreadsheet('2018B-2019A Telescope Schedules.xlsx')

        # Record the earliest date for each site: any date before this will return an empty set of Resources.
        # Record the latest date for each site: any date after this will return the Resources on this date.
        self._earliest_date_per_site = {site: min(self._resources[site], default=None) for site in self._sites}
        self._latest_date_per_site = {site: max(self._resources[site], default=None) for site in self._sites}
        for site in self._sites:
            # Only one of these checks should be necessary.
            if self._earliest_date_per_site is None or self._latest_date_per_site[site] is None:
                raise ValueError(f'No site resource data for {site.name}.')

    def _load_fpu_to_barcodes(self, site: Site, name: str) -> NoReturn:
        """
        FPUs at each site map to a unique barcode as defined in the files:
            * gmos[ns]_fpu_barcode.txt
        These are site-dependent values.
        """
        with open(os.path.join(self._path, name)) as f:
            for row in f:
                fpu, barcode = row.split()

                # Only map if the FPU is a resource.
                if fpu is not None:
                    self._itcd_fpu_to_barcode[site][fpu] = self._lookup_resource(barcode)

    def _load_csv(self, site: Site, name: str, c: Callable[[List[str]], Set[str]]) -> NoReturn:
        """
        Process a CSV file as a table, where:

        1. The first entry is a date in YYYY-mm-dd format
        2. The remaining entries are resources available on that date to the following date in the file.

        If a date is missing from the CSV file, copy the data from the previously defined date through to just before
        the new date.
        """
        with open(os.path.join(self._path, name)) as f:
            reader = csv.reader(f, delimiter=',')
            prev_row_date: Optional[date] = None

            for row in reader:
                row_date = datetime.strptime(row[0].strip(), '%Y-%m-%d').date()

                # Fill in any gaps by copying prev_row_date until we reach one less than row_date.
                if prev_row_date is not None:
                    missing_row_date = prev_row_date + ResourceMock._day
                    while missing_row_date < row_date:
                        # Make sure there is an entry and append to it to avoid overwriting anything already present.
                        date_set = self._resources[site].setdefault(missing_row_date, set())
                        self._resources[site][missing_row_date] = date_set | copy(self._resources[site][prev_row_date])
                        missing_row_date += ResourceMock._day

                # Get or create date_set for the date, and append new resources from table, ignoring blank entries.
                date_set = self._resources[site].setdefault(row_date, set())
                new_entries = {self._lookup_resource(r) for r in c(row[1:]) if r}
                self._resources[site][row_date] = date_set | new_entries

                # Advance the previous row date where data was defined.
                prev_row_date = row_date

    def _load_spreadsheet(self, name: str) -> NoReturn:
        """
        Process an Excel spreadsheet containing instrument, mode, and LGS information.

        The Excel spreadsheets have information available for every date, so we do not have to concern ourselves
        as in the _load_csv file above.
        """
        workbook = load_workbook(filename=os.path.join(self._path, name))
        for site in self._sites:
            sheet = workbook[site.name]

            # Read the sheet, skipping the header row (row 1).
            for row in sheet.iter_rows(min_row=2):
                # Skip shutdowns.
                mode = self._lookup_resource(row[1].value)
                if mode == 'Shutdown':
                    continue

                row_date = row[0].value.date()
                date_set = self._resources[site].setdefault(row_date, set())
                lgs = str_to_bool(row[2].value)

                # Filter out any ports that are empty.
                # Translate F2 to Flamingos2.
                instruments = {self._lookup_resource('Flamingos2' if c.value == 'F2' else c.value)
                               for c in row[3:] if c.value is not None}

                # TODO: Discuss with Bryan how to handle modes other than shutdowns?
                # TODO: It doesn't seem correct to include the mode in the Resource set.
                # date_set |= instruments | {mode}
                self._resources[site][row_date] = date_set | instruments
                self._lgs[site][row_date] = lgs

    def _lookup_resource(self, resource_id: str) -> Resource:
        """
        Function to perform Resource caching and minimize the number of Resource objects by attempting to reuse
        Resource objects with the same ID.

        Check if a Resource with id already exists.
        If it does, return it.
        If not, create it, add it to the map of all Resources, and then return it.

        Note that even if multiple objects do exist with the same ID, they will be considered equal by the
        Resource equality comparator.
        """
        # The Resource constructor raises an exception for id None or containing any capitalization of "none".
        if resource_id not in self._all_resources:
            self._all_resources[resource_id] = Resource(id=resource_id)
        return self._all_resources[resource_id]

    def date_range_for_site(self, site: Site) -> Tuple[date, date]:
        """
        Return the date range (inclusive) for which we have resource data for a site.
        """
        if site not in self._sites:
            raise ValueError(f'Request for resource dates for illegal site: {site.name}')
        return self._earliest_date_per_site[site], self._latest_date_per_site[site]

    def get_resources(self, site: Site, night_date: date) -> FrozenSet[Resource]:
        """
        For a site and a night date, return the set of available resources.
        If the date falls before any resource data for the site, return the empty set.
        If the date falls after any resource data for the site, return the last resource set.
        """
        if site not in self._sites:
            raise ValueError(f'Request for resources for illegal site: {site.name}')

        # If the date is before the first date, return an empty set.
        if night_date < self._earliest_date_per_site[site]:
            return frozenset()

        # If the date is past the last date, return the resources on the last date.
        actual_date = min(self._latest_date_per_site[site], night_date)
        return frozenset(self._resources[site][actual_date])

    def get_resources_for_sites(self,
                                sites: Collection[Site],
                                night_date: date) -> Dict[Site, FrozenSet[Resource]]:
        """
        For a collection of sites and a night date, return the set of available resources.
        """
        return {site: self.get_resources(site, night_date) for site in sites}

    def get_resources_for_dates(self,
                                site: Site,
                                night_dates: Collection[date]) -> Dict[date, FrozenSet[Resource]]:
        """
        For a site and a collection of night dates, return the set of available resources.
        """
        night_date_set = frozenset((d for d in night_dates))
        return {d: self.get_resources(site, d) for d in night_date_set}

    def get_resources_for_sites_and_dates(self,
                                          sites: Collection[Site],
                                          night_dates: Collection[date]) -> Dict[Site, Dict[date, FrozenSet[Resource]]]:
        """
        For a collection of sites and night dates, return the set of available resources.
        """
        site_set = frozenset((site for site in sites))
        return {site: self.get_resources_for_dates(site, night_dates) for site in site_set}

    def fpu_to_barcode(self, site: Site, fpu_name: str) -> Optional[Resource]:
        """
        Convert a long FPU name into the barcode, if it exists.
        """
        itcd_fpu_name = self._convert_fpu_to_itcd_name(site, fpu_name)
        return self._itcd_fpu_to_barcode[site].get(itcd_fpu_name)

    def _convert_fpu_to_itcd_name(self, site: Site, fpu_name: str) -> Optional[str]:
        """
        Convert a long FPU name into its ITCD name, if it exists.
        """
        if Site.GN in self._sites and site == Site.GN:
            return self._gmosn_ifu_dict.get(fpu_name)
        if Site.GS in self._sites and site == Site.GS:
            return self._gmoss_ifu_dict.get(fpu_name)
        return None

    def lookup_resource(self, resource_id: str) -> Optional[Resource]:
        """
        Given a resource ID, look it up and retrieve the Resource object from the cache if it exists.
        If not, None is returned.
        """
        return self._all_resources.get(resource_id)


# For Bryan and Kristin: testing instructions
if __name__ == '__main__':
    # To get the Resources for a specific site on a specific date, modify the following:
    st = Site.GN
    day = date(year=2018, month=11, day=8)

    resources_available = ResourceMock().get_resources(st, day)

    print(f'*** Resources for site {st.name} for {day} ***')
    for resource in sorted(resources_available, key=lambda x: x.id):
        print(resource)
    # print(', '.join([str(a) for a in sorted(resources_available, key=lambda x: x.id)]))
