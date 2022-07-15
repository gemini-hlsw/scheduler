import json
import os
from typing import NoReturn, Union, List

from astropy import units as u
from openpyxl import Workbook

from api.programprovider.abstract import ProgramProvider
from api.programprovider.ocs import OcsProgramProvider
from common.minimodel import Atom, Group, Observation, ObservationClass, Program
from common.plans import Plans
from components.collector import Collector, NightEventsManager


def print_program_from_provider(filename=os.path.join('data', 'GN-2018B-Q-101.json'),
                                provider: ProgramProvider = OcsProgramProvider) -> NoReturn:
    """
    Using a specified JSON file and a ProgramProvider, read in the program
    and print it.

    TODO: Could pass in JSON data instead, as GppProgramProvider will not produce files.
    """
    with open(filename, 'r') as f:
        data = json.loads(f.read())

    program = provider.parse_program(data['PROGRAM_BASIC'])
    print_program(program)


def print_program(program: Program) -> NoReturn:
    """
    Print the high-level information about a program in human semi-readable format
    to give an idea as to its structure.
    """
    print(f'Program: {program.id}')

    def sep(depth: int) -> str:
        return '----- ' * depth

    def print_observation(depth: int, obs: Observation) -> NoReturn:
        print(f'{sep(depth)} Observation: {obs.id}')
        for atom in obs.sequence:
            print(f'{sep(depth + 1)} {atom}')

    def print_group(depth: int, group: Group) -> NoReturn:
        # Is this a subgroup or an observation?
        if isinstance(group.children, Observation):
            print_observation(depth, group.children)
        elif isinstance(group.children, list):
            print(f'{sep(depth)} Group: {group.id}')
            for child in group.children:
                print_group(depth + 1, child)

    # Print the group and atom information.
    print_group(1, program.root_group)


def print_collector_info(collector: Collector, samples: int = 60) -> NoReturn:
    # Output some information.
    print(f'Pre-Collector / Collector running from:')
    print(f'   start time:       {collector.start_time}')
    print(f'   end time:         {collector.end_time}')
    print(f'   time slot length: {collector.time_slot_length.to(u.min)}')

    # Print out sampled calculation for every hour as there are far too many results to print in full.
    samples = 60
    time_grid = collector.time_grid
    for site in collector.sites:
        print(f'\n\n+++++ NIGHT EVENTS FOR {site.name} +++++')
        night_events = NightEventsManager.get_night_events(collector.time_grid, collector.time_slot_length, site)
        for idx, jday in enumerate(time_grid):
            start = night_events.local_times[idx][0]
            end = night_events.local_times[idx][-1]
            num_slots = len(night_events.times[idx])
            print(f'* DAY {idx}: {start} to {end}, {num_slots} time slots.')
            print(f'\tmidnight:         {night_events.midnight[idx]}')
            print(f'\tsunset:           {night_events.sunset[idx]}')
            print(f'\tsunrise:          {night_events.sunrise[idx]}')
            print(f'\t12° eve twilight: {night_events.twilight_evening_12[idx]}')
            print(f'\t12° mor twilight: {night_events.twilight_morning_12[idx]}')
            print(f'\tmoonrise:         {night_events.moonrise[idx]}')
            print(f'\tmoonset:          {night_events.moonset[idx]}')
            print(f'\n\tSun information (deg, sampled every {samples} time slots):')
            print(f'\tAlt:    {[a.to_value(u.deg) for a in night_events.sun_alt[idx][::samples]]}')
            print(f'\tAz:     {[a.to_value(u.deg) for a in night_events.sun_az[idx][::samples]]}')
            print(f'\tParAng: {[a.to_value(u.deg) for a in night_events.sun_par_ang[idx][::samples]]}')
            print(f'\n\tMoon information (km or deg, sampled every {samples} time slots):')
            print(f'\tDist:   {[a.to_value(u.km) for a in night_events.moon_dist[idx][::samples]]}')
            print(f'\tAlt:    {[a.to_value(u.deg) for a in night_events.moon_alt[idx][::samples]]}')
            print(f'\tAz:     {[a.to_value(u.deg) for a in night_events.moon_az[idx][::samples]]}')
            print(f'\tParAng: {[a.to_value(u.deg) for a in night_events.moon_par_ang[idx][::samples]]}')
            print(f'\n\tSun-Moon angle (deg, sampled every {samples} time slots):')
            print(f'\t{[a.to_value(u.deg) for a in night_events.sun_moon_ang[idx][::samples]]}')
            print('\n\n')

    target_info = sorted((obs_id, Collector.get_base_target(obs_id).name)
                         for obs_id in Collector.get_observation_ids())
    for obs_id, target_name in target_info:
        print(f'Observation {obs_id}, Target {target_name}')


def print_atoms_for_observation(observation: Observation) -> NoReturn:
    for atom in observation.sequence:
        print(f'\t{atom}')


def atoms_to_sheet(dt: Union[Program, Observation, Group]) -> NoReturn:
    """
    Print out the atoms in a program or observation to a spreadsheet.
    """

    wb = Workbook()
    ws = wb.active
    ws.append(['id', 'exec_time', 'prog_time', 'part_time', 'observed', 'qa_state', 'guide_state'])

    def save_to_sheet(atom: Atom):
        ws.cell(column=1, row=atom.id + 1, value=atom.id)
        ws.cell(column=2, row=atom.id + 1, value=atom.exec_time.total_seconds())
        ws.cell(column=3, row=atom.id + 1, value=atom.prog_time.total_seconds())
        ws.cell(column=4, row=atom.id + 1, value=atom.part_time.total_seconds())
        ws.cell(column=5, row=atom.id + 1, value=atom.observed)
        ws.cell(column=6, row=atom.id + 1, value=atom.qa_state.value)
        ws.cell(column=7, row=atom.id + 1, value=atom.guide_state)

    # TODO: Output for larger formats(e.g Program) not required but might be good to have.
    if isinstance(dt, Program):
        for obs in dt.observations():
            if obs.obs_class in [ObservationClass.SCIENCE, ObservationClass.PARTNERCAL]:
                for atom in obs.sequence:
                    ws.title = f'{obs.id}'
                    save_to_sheet(atom)
                ws = wb.create_sheet()
                ws.append(['id', 'exec_time', 'prog_time', 'part_time', 'observed', 'qa_state', 'guide_state'])
        wb.save(f'{dt.id}.xlsx')

    elif isinstance(dt, Observation):
        print(f'Observation: {dt.id}')
        for atom in dt.sequence:
            save_to_sheet(atom)
        wb.save(f'{dt.id}.xlsx')
    elif isinstance(dt, Group):
        raise NotImplementedError
    else:
        raise ValueError(f'Unsupported type: {type(dt)}')


def print_plans(all_plans: List[Plans]) -> NoReturn:
    """
    Print out the visit plans.
    """
    
    for plans in all_plans:
        print(f'\n\n+++++ NIGHT {plans.night + 1} +++++')
        for plan in plans:
            print(f'Plan for site: {plan.site.name}')
            for visit in plan.visits:
                print(f'\t{visit.start_time}   {visit.obs_id}')
