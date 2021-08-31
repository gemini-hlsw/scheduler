import csv
from os import getcwd

from openpyxl import load_workbook
from typing import Dict, List, NoReturn
from greedy_max import Site
from datetime import datetime, timedelta

from resource_mock.resources import Resources

class Resource:
    def __init__(self,path):
        self.path = getcwd()+path
        self.fpu = {}
        self.fpur = {}
        self.grat = {}
        self.fpu_to_barcode = {}
        self.instruments = {Site.GS: {}, Site.GN: {}}
        self.mode = {Site.GS: {}, Site.GN: {}}
        self.lgs = {Site.GS: {}, Site.GN: {}}
        self.ifu = {Site.GS: {}, Site.GN: {}}

    def _load_fpu(self, name: str, site: str) -> Dict[str,str]:
        barcodes = {}
        ifu = {}
        with open(f'{self.path}/GMOS{site.upper()}_{name}201789.txt') as f:
            reader =  csv.reader(f, delimiter=',') 
            for row in reader:
                barcodes[datetime.strptime(row[0].strip(),"%Y-%m-%d")] = [i.strip() for i in row[2:]]
                ifu[datetime.strptime(row[0].strip(),"%Y-%m-%d")] = row[1].strip()
        return ifu, barcodes
                
    def _load_grat(self, site: str) -> Dict[str,str]:
        out_dict = {}
        with open(f'{self.path}/GMOS{site.upper()}_GRAT201789.txt') as f:
            reader =  csv.reader(f, delimiter=',') 
            for row in reader:
                out_dict[datetime.strptime(row[0].strip(),"%Y-%m-%d")] = [i.strip().replace('+','') for i in row[1:]]
                out_dict[datetime.strptime(row[0].strip(),"%Y-%m-%d")].append('MIRROR') # Add mirror for GMOS 
        return out_dict

    def _load_f2b(self, site: str) -> Dict[str,str]:
        out_dict = {}
        with open(f'{self.path}/gmos{site}_fpu_barcode.txt') as f:
            
            for row in f:
                fpu, barcode = row.split()
                out_dict[fpu] = barcode
        return out_dict
    
    def _to_bool(self, b: str) -> bool:
        return b == 'Yes'
    
    def _nearest(self, items, pivot):
        return min(items, key=lambda x: abs(x - pivot))

    def _previous(self, items, pivot):
        # Return date equal or previous to pivot
        tdmin = timedelta.min
        tdzero = timedelta(days=0)
        result = None
        # result = min(items, key=lambda x: x)
        for item in items:
            diff = item - pivot
            if tdzero >= diff > tdmin:
                result = item
                dmin = diff
        return result

    def _excel_reader(self) -> NoReturn:

        sites = [site for site in Site]
        workbook = load_workbook(filename=f'{self.path}/2018B-2019A Telescope Schedules.xlsx')
        for site in sites:
            sheet = workbook[site.name]
            for row in sheet.iter_rows(min_row=2):                
                date = row[0].value
                self.instruments[site][date] = [c.value for c in row[3:]] 
                self.mode[site][date] = row[1].value
                self.lgs[site][date] = self._to_bool(row[2].value)
            
        if not self.instruments or not self.mode or not self.lgs:
            raise Exception("Problems on reading spreadsheet...") 

    def connect(self) -> NoReturn:
        """
        Allows the mock to load all the data locally, emulating a connection to the API.

        """
        print('Get Resource data...')
        sites = [site for site in Site]
        for site in sites:

            _site = 's' if site == Site.GS else 'n'
            self.ifu[site]['FPU'], self.fpu[site] = self._load_fpu('FPU',_site)
            self.ifu[site]['FPUr'], self.fpur[site] = self._load_fpu('FPUr',_site)
            self.grat[site] = self._load_grat(_site)
            self.fpu_to_barcode[site] = self._load_f2b(_site)

        if not self.fpu or not self.fpur or not self.grat:
            raise Exception("Problems on reading files...") 
        
        self._excel_reader()

    def _get_info(self, info: str, site: Site, date_str: str):
        
        date = datetime.strptime(date_str,"%Y-%m-%d")

        info_types = { 'fpu': self.fpu[site], 
                       'fpur': self.fpur[site],
                       'grat': self.grat[site],
                       'instr': self.instruments[site], 
                       'LGS': self.lgs[site], 
                       'mode': self.mode[site], 
                       'fpu-ifu': self.ifu[site]['FPU'], 
                       'fpur-ifu': self.ifu[site]['FPUr'] }

        if info in info_types:
            previous_date = self._previous(info_types[info].keys(), date)
            return info_types[info][previous_date]
            # if date in info_types[info]:
            #     return info_types[info][date]
            # else:
            #     nearest_date = self._previous(info_types[info].keys(), date)
            #     return info_types[info][nearest_date]
                
        else:
            print(f'No information about {info} is stored')
            return None

    def night_info(self, info_name: str, sites: List[Site], night_date: str):
        return {site: self._get_info(info_name, site, night_date) for site in sites }
    
    def get_night_resources(self, sites, night_date):
        fpu = self.night_info('fpu', sites, night_date)
        fpur = self.night_info('fpur', sites, night_date)
        grat = self.night_info('grat', sites, night_date)
        instruments = self.night_info('instr', sites, night_date)
        lgs = self.night_info('LGS', sites, night_date)
        modes = self.night_info('mode', sites, night_date)
        ifus = {'FPU':None, 'FPUr':None}
        ifus['FPU'] = self.night_info('fpu-ifu', sites, night_date)
        ifus['FPUr'] = self.night_info('fpur-ifu', sites, night_date)
        fpu2b = self.fpu_to_barcode

        return Resources(fpu,fpur,grat,instruments,lgs,modes,fpu2b,ifus)
